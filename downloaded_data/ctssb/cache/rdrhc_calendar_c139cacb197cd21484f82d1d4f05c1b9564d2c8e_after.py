from datetime import datetime, timedelta
from django.db.models import Q
import logging
import openpyxl
import re
from unipath import Path
import xlrd

# Setup logger
log = logging.getLogger(__name__)

# CLASSES
class RawShift(object):
    """Holds the details for a user's specified shift details"""
    
    def __init__(self, shift, date, comment):
        self.shift_code = shift
        self.start_date = date
        self.comment = comment

    def __str__(self):
        return "{} ({})".format(self.shift_code, self.start_date)

class FormattedShift(object):
    """Holds expanded details on a user's specified shift"""

    def __init__(self, code, start_datetime, end_datetime, comment, django):
        self.shift_code = code
        self.start_datetime = start_datetime
        self.end_datetime = end_datetime
        self.comment = comment
        self.django_shift = django

    def __str__(self):
        return "{} ({} to {})".format(
            self.shift_code, self.start_datetime, self.end_datetime
        )

class Schedule(object):
    """Holds all the users shifts and any noted modifications"""

    def __init__(
        self, shifts, additions, deletions, changes, missing, null, 
        missing_upload
    ):
        self.shifts = shifts
        self.additions = additions
        self.deletions = deletions
        self.changes = changes
        self.missing = missing
        self.null = null
        self.missing_upload = missing_upload

class EmailShift(object):
    """Holds details on shift modifications for emailing to the user"""
    date = 0
    msg = ""

    def __init__(self, date, msg):
        self.date = date
        self.msg = msg


def get_formatted_date(date):
    """Converts Python date object into string (as yyyy-mmm-dd)"""
    day = date.strftime("%d")
    
    month = date.strftime("%m")
    
    if month == "01":
        month = "JAN"
    elif month == "02":
        month = "FEB"
    elif month == "03":
        month = "MAR"
    elif month == "04":
        month = "APR"
    elif month == "05":
        month = "MAY"
    elif month == "06":
        month = "JUN"
    elif month == "07":
        month = "JUL"
    elif month == "08":
        month = "AUG"
    elif month == "09":
        month = "SEP"
    elif month == "10":
        month = "OCT"
    elif month == "11":
        month = "NOV"
    elif month == "12":
        month = "DEC"
    
    year = date.strftime("%Y")
    
    return ("{}-{}-{}".format(year, month, day))

def generate_calendar(user, schedule, cal_loc):
    """Generates an .ics file from the extracted user schedule"""
    
    log.info("Generating .ics calendar for %s" % user.name)

    # Generate initial calendar information
    lines = []
    
    lines.append("BEGIN:VCALENDAR")
    lines.append("PRODID:-//StudyBuffalo.com//RDRHC Calendar//EN")
    lines.append("VERSION:2.0")
    lines.append("CALSCALE:GREGORIAN")
    lines.append("X-WR-CALNAME:Work Schedule")
    lines.append("X-WR-TIMEZONE:America/Edmonton")
    lines.append("BEGIN:VTIMEZONE")
    lines.append("TZID:America/Edmonton")
    lines.append("X-LIC-LOCATION:America/Edmonton")
    lines.append("BEGIN:DAYLIGHT")
    lines.append("TZOFFSETFROM:-0700")
    lines.append("TZOFFSETTO:-0600")
    lines.append("TZNAME:MDT")
    lines.append("DTSTART:19700308T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=2SU")
    lines.append("END:DAYLIGHT")
    lines.append("BEGIN:STANDARD")
    lines.append("TZOFFSETFROM:-0600")
    lines.append("TZOFFSETTO:-0700")
    lines.append("TZNAME:MST")
    lines.append("DTSTART:19701101T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=11;BYDAY=1SU")
    lines.append("END:STANDARD")
    lines.append("END:VTIMEZONE")
    
	# Cycle through schedule and generate schedule events
    dt_stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    i = 0

    log.debug("Cycling through shifts for {}".format(user.name))

    for shift in schedule:
        try:
            start_date = shift.start_datetime.strftime("%Y%m%d")
            comment = shift.comment
        except Exception:
            log.error(
                "Unable to extract shift data to generate calendar", 
                exc_info=True
            )
            start_date = "20010101"
            comment = ""

        lines.append("BEGIN:VEVENT")

        if user.full_day == False:
            try:
                start_time = str(shift.start_datetime.time()).replace(":", "").zfill(6)
                end_date = shift.end_datetime.strftime("%Y%m%d")
                end_time = str(shift.end_datetime.time()).replace(":", "").zfill(6)
            except Exception:
                log.error(
                    "Unable to generate shift times for calendar",
                    exc_info=True
                )
                start_time = "000000"
                end_date = "20010102"
                end_time = "000000"

            lines.append("DTSTART;TZID=America/Edmonton:%sT%s" % (start_date, start_time))
            lines.append("DTEND;TZID=America/Edmonton:%sT%s" % (end_date, end_time))
        elif user.full_day == True:
            start_time = "000000"

            try:
                end_date = shift.start_datetime.date() + timedelta(days=1)
                end_date = end_date.strftime("%Y%m%d")
            except Exception:
                log.error(
                    "Unable to generate full day shift for schedule",
                    exc_info=True
                )
                end_date = "20010102"
                end_time = "000000"

            lines.append("DTSTART;VALUE=DATE:%s" % start_date)
            lines.append("DTEND;VALUE=DATE:%s" % end_date)
        
        lines.append("DTSTAMP:%s" % dt_stamp)
        lines.append("UID:%sT%s@studybuffalo.com-%s" % (start_date, start_time, i))
        lines.append("CREATED:%s" % dt_stamp)
        lines.append("DESCRIPTION:%s" % comment)
        lines.append("LAST-MODIFIED:%s" % dt_stamp)
        lines.append("LOCATION:Red Deer Regional Hospital Centre")
        lines.append("SEQUENCE:0")
        lines.append("STATUS:CONFIRMED")
        lines.append("SUMMARY:%s Shift" % shift.shift_code)
        lines.append("TRANSP:TRANSPARENT")

        if user.reminder != None:
            try:
                # Set the description text
                if user.reminder == 0:
                    alarm_description = (
                        "DESCRIPTION:{} shift starting now"
                    ).format(shift.shift_code)
                elif user.reminder == 1:
                    alarm_description = (
                        "DESCRIPTION:{} shift starting in {} minute"
                    ).format(shift.shift_code, user.reminder)
                else:
                    alarm_description = (
                        "DESCRIPTION:{} shift starting in {} minutes"
                    ).format(shift.shift_code, user.reminder)

                lines.append("BEGIN:VALARM")
                lines.append("TRIGGER:-PT{}M".format(user.reminder))
                lines.append("ACTION:DISPLAY")
                lines.append(alarm_description)
                lines.append("END:VALARM")
            except Exception:
                log.error("Unable to set reminder for shift", exc_info=True)

                lines.append("BEGIN:VALARM")
                lines.append("TRIGGER:-PT30M")
                lines.append("ACTION:DISPLAY")
                lines.append("DESCRIPTION:Shift starts in 30 minutes")
                lines.append("END:VALARM")
        lines.append("END:VEVENT")

        i = i + 1


    # End calendar file
    lines.append("END:VCALENDAR")

    # Fold any lines > 75 characters long
    log.debug("Folding lines greater than 75 characters long")

    folded = []
    
    for line in lines:
        if len(line) > 75:

            # Create line with first 75 characters
            newLine = line[0:75]
            folded.append(newLine + "\n")

            # Go through remainder and fold them
            line = line[75:]
            length = len(line)

            while length > 75:
                # Add folded line
                newLine = line[0:75]
                folded.append(" " + newLine + "\n")
                
                # Generate new line and length
                line = line[75:]
                length = len(line)

            # Add remainder
            folded.append(" " + line + "\n")
        else:
            folded.append(line + "\n")

    # Cycle through schedule list and generate .ics file
    calendar_name = user.calendar_name
    cal_title = "{}.ics".format(calendar_name)
    file_loc = Path(cal_loc, cal_title)

    log.debug("Saving calendar to {}".format(file_loc))

    with open(file_loc, "w") as ics:
        for line in folded:
            ics.write(line)

def extract_raw_schedule(book, sheet, user, index, row_start, row_end, date_col):
    """Returns an array of schedule_shift objects"""
    
    # EXTRACT SCHEDULE DETAILS FROM EXCEL DOCUMENT
    log.info("Extracting schedule details for %s" % user.name)

    # Generate comment map if this is an xls file
    if user.role == "a" or user.role == "t":
        comment_map = sheet.cell_note_map

    # Cycle through each row and extract shift date, code, and comments
    log.debug("Cycling through rows of excel schedule")

    shifts = []

    for i in range(row_start, row_end):
        # Extract date
        try:
            if user.role == "p":
                date = sheet.cell(row=i, column=date_col).value.date()
            elif user.role == "a" or user.role == "t":
                date = xlrd.xldate_as_tuple(
                    sheet.cell(i, date_col).value, book.datemode
                )
                date = datetime(*date).date()
        except AttributeError:
            # Expected error when there is no date value
            date = ""
        except IndexError:
            # Expected error when there is no date value
            date = ""
        except TypeError:
            # Expected error when there is no date value
            date = ""
        except Exception:
            if user.role == "p":
                sheet.cell(row=i, column=date_col).value
            elif user.role == "a" or user.role == "t":
                value = sheet.cell(i, date_col).value
            log.error(
                "Unable to extract date for user {} - row = {}, value = {}".format(user.name, i, value),
                exc_info=True
            )
            date = ""

		# Extract shift code
        try:
            if user.role == "p":
                shift_codes = sheet.cell(row=i, column=index).value.upper()
            elif user.role == "a" or user.role == "t":
                shift_codes = sheet.cell(i, index).value.upper()
        except AttributeError:
            # Expected error when there is no shift code value
            shift_codes = ""
        except IndexError:
            # Expect error when there is no shift code value
            shift_codes = ""
        except Exception:
            log.error(
                "Unable to extract shift code from worksheet in row {}".format(i),
                exc_info=True
            )
            shift_codes = ""

        # Extract cell comments
        comment = ""

        try:
            if user.role == "p":
                comment = sheet.cell(row=i, column=index).comment
            elif user.role == "a" or user.role == "t":
                comment = comment_map[i, index].text

            if comment is None:
                # Replaces "None" comments as empty string for calendar use
                comment = ""
            else:
                comment = str(comment)
                comment = comment.replace("\n", " ")
                comment = comment.strip()
        except KeyError:
            # Expected error when there is no comment
            comment = ""
        except Exception:
            log.error(
                "Unable to extract comments from worksheet in row {}".format(i),
                exc_info=True
            )
            comment = ""

        # Add shift to master list if it has a date and shift code
        if shift_codes != "" and re.match(r"^\s+$", shift_codes) and date != "":
            # Split each shift code on spaces or slashes
            shift_codes = re.split(r"(?:\s|/)+", shift_codes)
            
            for code in shift_codes:
                shifts.append(RawShift(code, date, comment))
                
                # Add pharmacist "X" shifts
                if user.role == "p" and code[-1:].upper() == "X":
                    shifts.append(RawShift("X", date, ""))
    
    # Sort the shifts by date
    # Note: should occur automatically, but just in case
    log.debug("Sorting shifts by date")

    sorted_shifts = sorted(shifts, key=lambda s: s.start_date)

    return sorted_shifts
    
def generate_formatted_schedule(user, raw_schedule, ShiftCode, StatHoliday, defaults, Shift):
    """Takes the raw schedule and returns the required formatted objects"""
    
    def collect_shift_codes(user, ShiftCode):
        """Takes a specific user and extracts the shift times"""
        log.debug("Collecting all the required shift codes for user")

        # Collect the user-specific codes
        user_codes = ShiftCode.objects.filter(
            Q(role=user.role) & Q(sb_user=user.sb_user)
        )

        # Collect the default codes (i.e. no user)
        default_codes = ShiftCode.objects.filter(
            Q(role=user.role) & Q(sb_user__isnull=True)
        )

        # Add all the user_codes into the codes list
        log.debug("Combining user-specific and default shift codes")

        codes = []
        
        for u_code in user_codes:
            codes.append(u_code)
        
        # Add any default codes that don't have a user code already
        for d_code in default_codes:
            if not any(d_code.code == code.code for code in codes):
                codes.append(d_code)
        
        return codes

    def collect_stat_holidays(schedule, StatHoliday):
        """Collects all stat holidays needed to generate a schedule"""
        try:
            first_day = schedule[0].start_date
            last_day = schedule[-1].start_date
        except Exception:
            log.warn(
                "Unable to retrieve statutory holidys based on schedule dates",
                exc_info=True
            )
            first_day = datetime(2001, 1, 1)
            last_day = datetime(2020, 12, 31)

        stat_holidays = StatHoliday.objects.all().filter(
            Q(date__gte=first_day) & Q(date__lte=last_day)
        )

        return stat_holidays

    def is_stat(stat_holidays, date):
        """Determines if the date is a stat holiday or not"""

        for holiday in stat_holidays:
            if date == holiday.date:
                return True
        
        return False

    def retrieve_old_schedule(user, Shift):
        """Retrieves the user's previous schedule from the database"""
        shifts = Shift.objects.all().filter(sb_user=user.sb_user).order_by("date")

        old_schedule = {}

        for shift in shifts:
            key_match = False
            shift_date = shift.date

            for key in old_schedule:
                if shift_date == key:
                    key_match = True

                    # Do not add "X" shifts
                    if shift.text_shift_code != "X":
                        # Append this shift to this key
                        old_schedule[shift_date].append(
                            RawShift(shift.text_shift_code, shift_date, "")
                        )

            if key_match == False:
                # Do not add "X" shifts
                if shift.text_shift_code != "X":
                    # Append a new key to the groupings
                    old_schedule[shift_date] = [
                        RawShift(shift.text_shift_code, shift_date, "")
                    ]

        return old_schedule

    def group_schedule_by_date(schedule):
        """Groups schedule shifts by date"""
        groupings = {}

        for shift in schedule:
            key_match = False
            shift_date = shift.start_datetime.date()

            for key in groupings:
                if shift_date == key:
                    key_match = True

                    # Do not add "X" shifts
                    if shift.shift_code != "X":
                        # Append this shift to this key
                        groupings[shift_date].append(
                            RawShift(shift.shift_code, shift_date, "")
                        )

            if key_match == False:
                # Do not add "X" shifts
                if shift.shift_code != "X":
                    # Append a new key to the groupings
                    groupings[shift_date] = [
                        RawShift(shift.shift_code, shift_date, "")
                    ]

        return groupings

    # Get shift codes/times for user
    shift_code_list = collect_shift_codes(user, ShiftCode)
    
    # Get all the stat holidays for the date range of the raw_schedule
    stat_holidays = collect_stat_holidays(raw_schedule, StatHoliday)
    
    # Assign start and end date/times to user's shifts
    schedule = []
    null_shifts = []
    missing_shifts = []
    missing_codes_for_upload = set()
    
    for shift in raw_schedule:
        # Search for a shift match
        shift_match = False
        
        # Record the day of the week
        try:
            dow = shift.start_date.weekday()
        except Exception:
            log.error("Unable to determine day of week", exc_info=True)
            dow = 0

        # Check if this is a stat holiday
        try:
            stat_match = is_stat(stat_holidays, shift.start_date)
        except Exception:
            log.error(
                "Unable to determine if this is a stat holiday", 
                exc_info=True
            )
            stat_match = False

        for code in shift_code_list:
	        # If matched, find the proper day to base shift details on
            
            if shift.shift_code == code.code:
                shift_match = True

                # Apply proper start time and duration
                if stat_match:
                    start_time = code.stat_start
                    duration = code.stat_duration
                elif dow == 0:
                    start_time = code.monday_start
                    duration = code.monday_duration
                elif dow == 1:
                    start_time = code.tuesday_start
                    duration = code.tuesday_duration
                elif dow == 2:
                    start_time = code.wednesday_start
                    duration = code.wednesday_duration
                elif dow == 3:
                    start_time = code.thursday_start
                    duration = code.thursday_duration
                elif dow == 4:
                    start_time = code.friday_start
                    duration = code.friday_duration
                elif dow == 5:
                    start_time = code.saturday_start
                    duration = code.saturday_duration
                elif dow == 6:
                    start_time = code.sunday_start
                    duration = code.sunday_duration

                if start_time:
                    # Shift has time, process as normal
                    
                    # Convert the decimal hours duration to h, m, and s
                    hours = int(duration)
                    minutes = int((duration*60) % 60)
                
                    start_datetime = datetime.combine(shift.start_date, start_time)

                    end_datetime = start_datetime + timedelta(
                        hours=hours,
                        minutes=minutes
                    )
  
                    schedule.append(FormattedShift(
                        shift.shift_code, start_datetime, end_datetime, 
                        shift.comment, code
                    ))
                else:
                    # Shift has no times - don't add to schedule and mark 
                    # it in the null shift list
                    msg = "{} - {}".format(
                        get_formatted_date(shift.start_date), shift.shift_code
                    )

                    null_shifts.append(
                        EmailShift(shift.start_date, msg)
                    )

                # End loop
                break

        # If no shift match, provide default values
        if shift_match == False:
            # Add missing shift to the Missing shift list
            msg = "{} - {}".format(
                get_formatted_date(shift.start_date), shift.shift_code
            )

            missing_shifts.append(
                EmailShift(shift.start_date, msg)
            )

            # Add the missing code to the missing code set
            missing_codes_for_upload.add(shift.shift_code)
                
            # Set default times
            if stat_match:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["stat_start"]
                )
                    
                end_datetime = start_datetime + timedelta(
                    hours=defaults["stat_hours"],
                    minutes=defaults["stat_minutes"]
                )
            elif dow >= 5:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["weekend_start"]
                )

                end_datetime = start_datetime + timedelta(
                    hours=defaults["weekend_hours"],
                    minutes=defaults["weekend_minutes"]
                )
            else:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["weekday_start"]
                )
                    
                end_datetime = start_datetime + timedelta(
                    hours=defaults["weekday_hours"],
                    minutes=defaults["weekday_minutes"]
                )
  
            schedule.append(FormattedShift(
                shift.shift_code, start_datetime, end_datetime, 
                shift.comment, None
            ))

    
    # Determine the shift additions, deletions, and changes
    # Retrieve the old schedule
    old_schedule = retrieve_old_schedule(user, Shift)
    
    # Generate a new schedule listing organized by date
    new_by_date = group_schedule_by_date(schedule)
    
    # Check if there are any deletions or changes
    deletions = []
    changes = []

    for old_date, old_shifts in old_schedule.items():
        shift_match = []

        if old_date in new_by_date:
            new_shifts = new_by_date[old_date]
            
            if len(old_shifts) == len(new_shifts):
                for old in old_shifts:
                    for new in new_shifts:
                        if old.shift_code == new.shift_code:
                            shift_match.append(True)

            # If the number of Trues equal length of old_shifts, 
            # no changes occurred
            if len(shift_match) != len(old_shifts):
                old_codes = "/".join(str(s.shift_code) for s in old_shifts)
                new_codes = "/".join(str(s.shift_code) for s in new_shifts)
                msg = "{} - {} changed to {}".format(
                   get_formatted_date(old_date), 
                   old_codes,
                   new_codes
                )

                changes.append(EmailShift(old_date, msg))
        else:
            # Shift was deleted
            old_codes = "/".join(str(s.shift_code) for s in old_shifts)
            msg = "{} - {}".format(
                get_formatted_date(old_date), 
                old_codes
            )

            deletions.append(EmailShift(old_date, msg))

    # Checks if there are any new additions
    additions = []

    for new_date, new_shifts in new_by_date.items():
        shift_match = []

        if new_date not in old_schedule:
            new_codes = "/".join(str(s.shift_code) for s in new_shifts)
            msg = "{} - {}".format(
                get_formatted_date(new_date), 
                new_codes
            )

            additions.append(EmailShift(new_date, msg))
    
    # Removes missing shifts not in the additions or changes lists
    # (user will have already been notified on these shifts)
    missing = []

    for m in missing_shifts:
        for c in changes:
            if m.date == c.date:
                missing.append(m)

        for a in additions:
            if m.date == a.date:
                missing.append(m)

    null = []

    # Removes null shifts not in the additions or changes lists
    # (user will have already been notified on these shifts)
    for n in null_shifts:
        for c in changes:
            if n.date == c.date:
                null.append(n)

        for a in additions:
            if n.date == a.date:
                null.append(n)

    # Return all the required items to generate the calendar and emails
    return Schedule(
        schedule, additions, deletions, changes, missing, null, 
        missing_codes_for_upload
    )

def return_column_index(sheet, user, name_row, col_start, col_end):
    """Determines the Excel column containing the provided user"""
    log.debug("Looking for user index in Excel schedule")

    index = None

    for i in range(col_start, col_end):
        try:
            if user.role == "p":
                cell_name = str(sheet.cell(row=name_row, column=i).value).strip()
            elif user.role == "a" or user.role == "t":
                cell_name = str(sheet.cell(name_row, i).value).strip()
            
            if cell_name.upper() == user.schedule_name.upper():
                index = i
                break
        except IndexError:
            # Expected error if loop exceeds Excel content
            break
        except Exception:
            log.critical(
                "Error while searching for column index for {}".format(user.name),
                exc_info=True
            )

    return index

def assemble_schedule(app_config, excel_files, user, ShiftCode, StatHoliday, Shift):
    # Setup the required Excel details
    role = user.role

    file_loc = excel_files[role]
    config = app_config["{}_excel".format(user.role).lower()]

    # Open the proper Excel worksheet
    log.debug("Opening the Excel worksheet")

    if user.role == "p":
        try:
            excel_book = openpyxl.load_workbook(file_loc)
            excel_sheet = excel_book[config["sheet"]]
        except Exception:
            log.critical(
                "Unable to open .xlsx file for user role = {}".format(user.role),
                exc_info=True
            )
    elif user.role == "a" or user.role == "t":
        try:
            excel_book = xlrd.open_workbook(file_loc)
            excel_sheet = excel_book.sheet_by_name(config["sheet"])
        except Exception:
            log.critical(
                "Unable to open .xls file for user role = {}".format(user.role),
                exc_info=True
            )

    # Find column index for this user
    user_index = return_column_index(
        excel_sheet, user, config["name_row"], config["col_start"], config["col_end"]
    )

    # If the user.index is found, can run rest of program
    if user_index:
        raw_schedule = extract_raw_schedule(
            excel_book, excel_sheet, user, user_index, 
            config["row_start"], config["row_end"], config["date_col"]
        )

        formatted_schedule = generate_formatted_schedule(
            user, raw_schedule, ShiftCode, StatHoliday, app_config["calendar_defaults"], Shift
        )

        return formatted_schedule
    else:
        log.warn(
            "Unable to find {} (role = {}) in the Excel schedule".format(
                user.name,
                user.role
            ),
        )

        return None