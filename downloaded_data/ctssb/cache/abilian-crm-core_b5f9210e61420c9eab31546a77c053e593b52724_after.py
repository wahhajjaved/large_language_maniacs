# coding=utf-8
"""
"""
from __future__ import absolute_import

import logging
import hashlib
import math
from functools import partial
from itertools import izip, islice, chain
import datetime
from operator import attrgetter
from collections import defaultdict
import pprint

import sqlalchemy as sa
import itsdangerous
import openpyxl
from openpyxl import Workbook, styles
from openpyxl.cell import STRING_TYPES
from openpyxl.utils import units, get_column_letter
from openpyxl.writer.dump_worksheet import WriteOnlyCell

from flask import current_app

from abilian.i18n import _
from abilian.core.extensions import db
from abilian.core.sqlalchemy import JSON as JSONType
from abilian.web.forms.fields import ModelFieldList
from abilian.services.vocabularies.models import BaseVocabulary

from ..models import PostalAddress
from .exc import ExcelError, ExcelImportError
from .columns import (
  Column, ColumnSet, RelatedColumnSet, ManyRelatedColumnSet,
  VocabularyColumn, PostalAddressColumn, DateColumn, DateTimeColumn,
  Invalid,
)


logger = logging.getLogger(__name__)


class ExcelManager(object):
  """
  import/export data to excel files, from/to model
  """

  XF_HEADER = {
    'font': styles.Font(bold=True),
    'alignment': styles.Alignment(horizontal='center', vertical='center',
                                  wrapText=True)
  }
  XF_EDITABLE = styles.Protection(locked=False)
  XF_DATE_EDITABLE = styles.Protection(locked=False)
  XF_DATE_NUMBER_FORMAT = "DD/MM/YYYY"

  #: column to identify item by (unique) name (if no database id is
  # available). This is used by importer to import a many related sheet
  ID_BY_NAME_COL = None

  #: list of columns that identify a unique object. If an import sheet has no id
  # one of this columns may identify an object. In this case the object is
  # updated with values from non-empty cells (like a dict.update())
  UNIQUE_ID_COLS = ()

  #: sheet name where data should be, at import or export
  MAIN_SHEET_NAME = u'Sheet 1'

  #: For many related export/import, set column at which object's columns is
  # splitted to insert related object columns
  MANY_SPLIT_COLUMN = 2

  #: list of field to skip
  SKIP_COLS = ()

  def __init__(self, model_cls, form_cls, many_related_cs=()):
    """
    model_cls: Model class
    form_cls: Form class
    many_related_cs: iterable of ManyRelatedColumnSet instances
    """
    self.MANY_RELATED_CS = many_related_cs
    self.model_cls = model_cls
    self.form_cls = form_cls
    self.form = form_cls(csrf_enabled=False)
    self.mapper = sa.inspect(model_cls)
    self.db_columns = self.mapper.c

    if self.ID_BY_NAME_COL and self.ID_BY_NAME_COL not in self.UNIQUE_ID_COLS:
      self.UNIQUE_ID_COLS += (self.ID_BY_NAME_COL,)

    # create signer
    config = current_app.config
    secret_key = config.get('SECRET_KEY')
    self.signer = itsdangerous.TimestampSigner(secret_key, salt=__name__)

    # collect exportable attributes
    columns = [Column('id', 'id', int)]

    for field in self.form:
      info = self.columns_for(field, self.form)
      if info is not None:
        columns.append(info)

    columns.extend(self.additional_columns())
    self.columns = ColumnSet(*columns)
    # logger.debug('\nColumns:\n %s', pprint.pformat([c for c in self.columns]))

  def additional_columns(self):
    """
    To be overriden by subclasses that want to add specific columns
    """
    return []

  @property
  def attrs(self):
    return self.columns.attrs

  def attrs_signature(self, columns=None):
    """
    Return hash for current labels.
    """
    if columns is None:
      columns = self.columns
    md5 = hashlib.md5()
    for attr in columns.attrs:
      self.update_md5(md5, attr)
    return md5.hexdigest()

  @property
  def labels(self):
    return self.columns.labels

  def get_attr_to_column(self, columnset, map_related_attr=False):
    """
    map attribute names to their respective Column instance.

    If map_related_attr is True, also map related's attributes to their
    RelatedColumnSet instance
    """
    attr_map = dict(izip(columnset.attrs, columnset.iter_flatened()))
    # we also want to collect related
    # attributes of RelatedColumnSets. self.columns.attrs doesn't show them
    for col in columnset:
      if isinstance(col, RelatedColumnSet):
        attr_map[col.related_attr] = col
        if map_related_attr:
          for attr in col.attrs:
            attr_map[attr] = col

    return attr_map

  @property
  def attr_to_column(self):
    if not hasattr(self, '_attr_to_column'):
      self._attr_to_column = self.get_attr_to_column(self.columns)
    return self._attr_to_column

  @property
  def attr_to_main_column(self):
    """ Return a mapping of attribute name: column instance, with the exception
    of attributes of related instances mapped to a RelatedColumnSet() instance.
    """
    if not hasattr(self, '_attr_to_main_column'):
      self._attr_to_main_column = self.get_attr_to_column(self.columns,
                                                          map_related_attr=True)
    return self._attr_to_main_column


  def export(self, objects, related_column_set=None, progress_callback=None):
    """
    Exports objects to a Workbook
    """
    if related_column_set is not None:
      return self.export_many(objects, related_column_set,
                              progress_callback=progress_callback)

    wb = Workbook()
    if wb.worksheets:
      wb.remove_sheet(wb.active)
    ws, row = self._new_export_sheet(wb, self.model_cls.__name__, self.columns)

    total = objects.count() if hasattr(objects, 'count') else len(objects)
    exported = 0
    if progress_callback:
      progress_callback(exported=exported, total=total)

    for r, obj in enumerate(objects, row):
      md5 = hashlib.md5()
      offset = 0
      cells = [WriteOnlyCell(ws)]
      for c, col in enumerate(self.columns, 1):
        # one model column might be exported in multiple excel columns. Example:
        # a related entity like a contact: fullname, email, etc
        for rel_idx, (import_val, value) in enumerate(col.data(obj)):
          dest_col = c + offset + rel_idx
          cell = WriteOnlyCell(ws, value=import_val)
          self.style_for(cell)
          cells.append(cell)
          self.update_md5(md5, import_val)

        offset += rel_idx

      cells[0].value=self.signer.sign(md5.hexdigest())
      ws.append(cells)

      exported += 1
      if progress_callback:
        progress_callback(exported=exported, total=total)

    if progress_callback:
        progress_callback(exported=total, total=total)

    self.finalize_worksheet(ws)
    return wb

  def export_many(self, objects, related_columns_set, progress_callback=None):
    """
    :param related_columns_set: a :class:`ManyRelatedColumnSet` instance
    """
    assert isinstance(related_columns_set, ManyRelatedColumnSet)

    wb = Workbook()
    if wb.worksheets:
      wb.remove_sheet(wb.active)

    total = objects.count() if hasattr(objects, 'count') else len(objects)
    exported = 0
    if progress_callback:
      progress_callback(exported=exported, total=total)

    all_columns = self._columns_for_many_related(related_columns_set)
    ws, start_row = self._new_export_sheet(wb,
                                           related_columns_set.export_label,
                                           all_columns)
    related_columns_len = related_columns_set.colspan

    row_offset = 0
    for r, obj in enumerate(objects, start_row):
      row_offset -= 1
      # prepare main item data: it will be repeated for each of its related
      # objects
      columns = iter(self.columns)

      head_data = [import_val
                   for col in islice(columns, self.MANY_SPLIT_COLUMN)
                   for import_val, value in col.data(obj)]

      # islice has started to consume 'columns' iterator, iter remaining columns
      tail_data = [import_val
                   for col in columns
                   for import_val, value in col.data(obj)]

      # generate rows for each related object
      is_empty = True
      md5 = hashlib.md5()
      for item in related_columns_set.iter_items(obj):
        cells = [WriteOnlyCell(ws)]
        is_empty = False
        row_offset += 1
        col_offset = 1
        for c, val in enumerate(head_data, 1):
          # FIXME: we could keep HEAD_MD5, and just extends `cells` with
          # head_data
          cell = WriteOnlyCell(ws, value=val)
          self.style_for(cell)
          cells.append(cell)
          self.update_md5(md5, value)
          # ws.write(r+row_offset, c, value, style)
          col_offset += 1

        data = related_columns_set.data(item)
        for c, (import_val, value) in enumerate(data, col_offset):
          cell = WriteOnlyCell(ws, value=import_val)
          self.style_for(cell)
          cells.append(cell)
          #ws.write(r+row_offset, c, import_val, self.style_for(import_val))
          self.update_md5(md5, import_val)
          col_offset += 1

        for c, val in enumerate(tail_data, col_offset):
          cell = WriteOnlyCell(ws, value=val)
          self.style_for(cell)
          cells.append(cell)
          #ws.write(r+row_offset, c, value, style)
          self.update_md5(md5, value)

        cells[0].value = self.signer.sign(md5.hexdigest())
        ws.append(cells)

      if is_empty:
        # no related items. We still need to output 1 line, with empty columns
        # for the related part.
        cells = [WriteOnlyCell(ws)]
        row_offset += 1
        col_offset = 1
        # head data 1st, then adjust col_offset to skip related item columns,
        # then tail data
        for c, cell in enumerate(head_data, 1):
          cells.append(cell)
          self.update_md5(md5, value)
          col_offset += 1

        for ignored in range(related_columns_len):
          cells.append(WriteOnlyCell(ws))

        col_offset += related_columns_len

        for c, cell in enumerate(tail_data, col_offset):
          cells.append(cell)
          self.update_md5(md5, value)

        cells[0].value = self.signer.sign(md5.hexdigest())
        ws.append(cells)

      exported += 1
      if progress_callback:
        progress_callback(exported=exported, total=total)

    if progress_callback:
        progress_callback(exported=total, total=total)

    self.finalize_worksheet(ws)
    return wb

  def _new_export_sheet(self, wb, name, columns):
    ws = wb.create_sheet(title=name)
    #ws.protect = True

    # attributes row
    row = 0
    md5 = hashlib.md5()
    cells = [WriteOnlyCell(ws)]
    for c, attr in enumerate(columns.attrs, 1):
      cells.append(WriteOnlyCell(ws, value=attr))
      self.update_md5(md5, attr)

    cells[0].value = self.signer.sign(md5.hexdigest())
    ws.append(cells)
    row +=1

    #  labels row
    md5 = hashlib.md5()
    cells = [WriteOnlyCell(ws)]
    for c, label in enumerate(columns.labels, 1):
      cell = WriteOnlyCell(ws, value=label)
      cell.font = self.XF_HEADER['font']
      cell.alignment = self.XF_HEADER['alignment']
      cells.append(cell)
      self.update_md5(md5, label)

    cells[0].value = self.signer.sign(md5.hexdigest())
    ws.append(cells)
    row += 1

    # row / column properties
    ws.row_dimensions[1].hidden = True #  hide attributes names row
    ws.column_dimensions['A'].hidden = True # hide md5 column
    MIN_WIDTH = units.BASE_COL_WIDTH
    MAX_WIDTH = MIN_WIDTH * 2
    overflow = 0

    for idx, cell in enumerate(cells, 1):
      letter = get_column_letter(idx)
      width = len(cell.value) + 1
      if width > MAX_WIDTH:
        overflow = max(overflow, width)
        width = MAX_WIDTH
      # BASE_COL_WIDTH <= custom width <= BASE_COL_WIDTH * 2
      ws.column_dimensions[letter].width = max(MIN_WIDTH, width)

    if overflow:
      lines = int(math.ceil(overflow / float(MAX_WIDTH)))
      ws.row_dimensions[2].height = units.DEFAULT_ROW_HEIGHT * lines

    return ws, row

  def finalize_worksheet(self, ws):
    # if we do this during initialize, next ws.append() will occur at line 4,
    # leaving a blank line.
    ws.freeze_panes = ws.cell(row=3, column=3)

  def _columns_for_many_related(self, related_cs):
    """
    Given a ManyRelatedColumnSet instance, returns a columns set with object
    columns and related columns
    """
    SPLIT_COLUMN = self.MANY_SPLIT_COLUMN
    main_cols = list(self.columns)
    head_cols = main_cols[:SPLIT_COLUMN]
    tail_cols = main_cols[SPLIT_COLUMN:]
    all_columns = ColumnSet(*list(chain(head_cols, (related_cs,), tail_cols)))
    return all_columns

  def style_for(self, cell):
    """
    Return XLWT style to use for value
    """
    cell.protection = self.XF_EDITABLE
    if isinstance(cell.value, datetime.date):
      cell.protection = self.XF_DATE_EDITABLE
      cell.number_format = self.XF_DATE_NUMBER_FORMAT

  def import_data(self, xls_file, many_related_columns_set=()):
    """ read xls file, detect changed rows; then for each find changed
    attributes of related item.

    return a list of rows; each row is a dict with:
      - `item`: db object
      - `modified`: dict attr: (current value, new value)
      - `attr_sig`: signature of 'modified', required for 'save_data()'
    """
    wb = openpyxl.load_workbook(xls_file)
    try:
      ws = wb[self.model_cls.__name__]
    except KeyError:
      # last chance
      ws = wb['Sheet 1']

    self._validate_import_sheet(ws)
    data_rows = self._collect_changed_rows(ws, wb)
    by_id = {}
    by_name = {}
    unique_id_cols = {}
    for d in data_rows:
      item_id = d.get('id')
      if item_id and not isinstance(item_id, Invalid):
        by_id[item_id] = d
      name = d['__metadata__'].get('unique_name')
      if name is not None:
        # FIXME: check name is unique
        by_name[name] = d
      uid_cols = d['__metadata__'].get('unique_id_cols')
      if uid_cols is not None:
        for key, val in uid_cols.iteritems():
          unique_id_cols.setdefault(key, {})[val] = d

    # collect data from the 'many related' sets
    for related_cs in many_related_columns_set:
      self._collect_changed_relateds(related_cs, wb, data_rows, by_id, by_name)

    # prefetch items
    ids = [d['id'] for d in data_rows if 'id' in d]
    q = self.model_cls.query
    q.filter(self.model_cls.id.in_(ids)).all()

    # detect modified attributes
    attr_to_col = self.attr_to_column
    modified_items = []

    for data in data_rows:
      item = None
      metadata = data.pop('__metadata__')
      is_new = metadata.get('is_new')
      is_update = metadata.get('is_update')

      if not is_new:
        try:
          item = q.get(data['id'])
        except:
          # this may be a deleted line from an exported file, but excel has
          # preserved hidden cell
          continue
        if item is None:
          continue
      else:
        # new item, or update sheet: incomplete lines of existing items
        item = self.model_cls()

        for col_name in unique_id_cols:
          val = data.get(col_name)
          if val is not None:
            try:
              item = q.filter(getattr(self.model_cls, col_name) == val).one()
            except sa.orm.exc.NoResultFound:
              pass
            except sa.orm.exc.MultipleResultsFound:
              # log this to sentry
              logger.error('Multiple items match %s[%s] == ""%s"',
                           self.model_cls.__name__, col_name,
                           unicode(val).encode('utf-8'),
                           extra={'stack': True,})
            else:
              is_new = metadata['is_new'] = False
              is_update = metadata['is_update'] = True
              data['id'] = item.id

      assert not is_update or (is_update and not is_new)

      many_related_data = {}
      for cs in self.MANY_RELATED_CS:
        many_related_data[cs.related_attr] = data.pop(cs.related_attr, None)

      modified = self._get_modified_data(item, data, attr_to_col, is_new,
                                         is_update)
      required_missing = modified.pop('__required_missing__', False)

      modified_relateds = {}
      for cs in self.MANY_RELATED_CS:
        related_data = many_related_data.get(cs.related_attr)
        if not related_data:
          continue

        rel_items = []
        name_getter = attrgetter(cs.ID_BY_NAME_COL)
        current_data = getattr(item, cs.related_attr)
        current_data_map = { name_getter(o): o for o in current_data }
        related_attr_to_col = self.get_attr_to_column(related_cs,
                                                      map_related_attr=True)

        for item_data in related_data:
          rel_metadata = item_data.pop('__metadata__')
          name = rel_metadata.get('__metadata__', {}).get('unique_name')
          is_new = name is not None
          default = cs.model_cls()
          obj = current_data_map.get(name, default) if is_new else default
          rel_modified = self._get_modified_data(
            obj, item_data, related_attr_to_col, is_new
          )

          if rel_modified:
            rel_items.append(dict(modified=rel_modified,
                                  metadata=rel_metadata))

        if rel_items:
          modified_relateds[cs.related_attr] = rel_items

      if modified or modified_relateds:
        # FIXME: use itsdangerous.TimedSerializer instead of joining a list
        valid_keys = (k for k in modified
                      if not (required_missing or modified[k].error))
        attr_sig = self.signer.sign(u';'.join(sorted(valid_keys)))
        attr_sig = self.extract_signature(attr_sig)

        modified_items.append(dict(item=item,
                                   metadata=metadata,
                                   required_missing=required_missing,
                                   modified=modified,
                                   many_relateds=modified_relateds,
                                   attr_signature=attr_sig))

    # FIXME: also return list of invalid_rows?
    return modified_items

  def _validate_import_sheet(self, ws, columns=None):
    """
    Validates signatures of column lines: columns ordering and labels should
    not have been changed
    """
    orig_md5 = self.get_cell_value(ws, 0, 0)
    try:
      orig_md5 = self.signer.unsign(orig_md5)
    except itsdangerous.BadSignature:
      raise ExcelError(u'La signature de la 1ere ligne est incorrecte, '
                       'impossible de valider le fichier')

    if columns is None:
      columns = self.columns

    md5 = hashlib.md5()
    for col in range(1, ws.ncols):
      attr = self.get_cell_value(ws, 0, col)
      self.update_md5(md5, attr)

    if orig_md5 != md5.hexdigest():
      raise ExcelError(u'L\'ordre des colonnes a changé ou leurs étiquettes ont'
                       u' été modifiées: le fichier n\'est pas valide.')

    if orig_md5 != self.attrs_signature(columns):
      raise ExcelError(
        u'La feuille "{}" utilise une version du format d\'export qui'
        u' n\'est plus à jour'.format(ws.name))


  def _collect_changed_rows(self, ws, wb):
    """
    collect modified rows: compare md5@column(0) with md5(row values), and
    new rows
    """
    attr_to_main = self.attr_to_main_column
    attr_to_col = self.attr_to_column
    single_value_attrs = {k for k, column in attr_to_main.items()
                          if not isinstance(column, RelatedColumnSet)}
    invalid_lines = []
    data_rows = []

    for row in range(2, ws.nrows):
      cell = partial(self.get_cell_value, ws, row)
      data = defaultdict(dict)
      orig_md5 = cell(0)
      is_new = False if orig_md5 else True
      data['__metadata__'] = {'row': row + 1,
                              'is_new': is_new,
                              'is_update': False,
                              'unique_id_cols': dict(),}

      if not is_new:
        try:
          orig_md5 = self.signer.unsign(orig_md5)
        except itsdangerous.BadSignature:
          invalid_lines.append(row)
          logger.debug('Skip line %d: md5 signature is invalid', row)
          continue

      md5 = hashlib.md5()
      for idx, attr in enumerate(self.attrs, 1):
        value = cell(idx)
        if is_new and attr == 'id':
          data['__metadata__']['id'] = value
          continue

        value = cell(idx)
        value = self.update_import_data(data, attr, value,
                                        attr_to_col, attr_to_main,
                                        self.ID_BY_NAME_COL,
                                        self.UNIQUE_ID_COLS,
                                        wb)
        self.update_md5(md5, value)

      md5 = md5.hexdigest()
      if is_new or orig_md5 != md5:
        for attr in single_value_attrs:
          if attr in data:
            data[attr] = data[attr][attr]
        data_rows.append(data)

    return data_rows

  def _collect_changed_relateds(self, related_cs, wb, data_rows, id_map,
                                name_map):
    """
    Collect changes from a many related column set. Data must be in sheet
    with named after `related_cs.export_label
    """
    if not self.ID_BY_NAME_COL:
      # the manager is not configured to be able find model by a name
      # column. Since main object id is not exported in 'many_related' exports
      # we cannot process imports for this model_cls.
      return

    try:
      ws = wb[related_cs.export_label]
    except KeyError:
      logger.debug('No sheet named "%s"',
                   related_cs.export_label.encode('utf-8'))
      return

    all_columns = self._columns_for_many_related(related_cs)
    try:
      self._validate_import_sheet(ws, all_columns)
    except ExcelError:
      logger.warning(
        'Many related _collect_changed_relateds: Invalid header signature, '
        'skipping (%s)', related_cs.export_label.encode('utf-8')
      )
      return

    attr_to_main = self.attr_to_main_column
    attr_to_col = self.attr_to_column
    related_attr_to_main = self.get_attr_to_column(related_cs, True)
    related_attr_to_col = self.get_attr_to_column(related_cs)

    single_value_attrs = {k for k, column in related_attr_to_main.items()
                          if not isinstance(column, RelatedColumnSet)}
    invalid_lines = []

    for row in range(2, ws.nrows):
      cell = partial(self.get_cell_value, ws, row)
      data = defaultdict(dict)
      data['__metadata__'] = {'row': row + 1}
      related_data = defaultdict(dict)
      related_data['__metadata__'] = {'row': row + 1}

      orig_md5 = cell(0)
      is_new = False if orig_md5 else True

      if not is_new:
        try:
          orig_md5 = self.signer.unsign(orig_md5)
        except itsdangerous.BadSignature:
          invalid_lines.append(row)
          logger.debug('Skip line %d: md5 signature is invalid', row)
          continue

      attrs = iter(self.attrs)
      md5 = hashlib.md5()
      # main item: head data
      for idx, attr in enumerate(islice(attrs, self.MANY_SPLIT_COLUMN), 1):
        if is_new and attr == 'id':
          continue
        value = cell(idx)
        value = self.update_import_data(data, attr, value,
                                        attr_to_col, attr_to_main,
                                        self.ID_BY_NAME_COL,
                                        wb)
        self.update_md5(md5, value)

      # related object data
      for idx, attr in enumerate(related_cs.attrs, idx + 1):
        value = cell(idx)
        value = self.update_import_data(
          related_data, attr, value,
          related_attr_to_col, related_attr_to_main,
          related_cs.ID_BY_NAME_COL,
          wb)
        self.update_md5(md5, value)

      # main item: tail data
      for idx, attr in enumerate(attrs, idx + 1):
        if is_new and attr == 'id':
          continue
        value = cell(idx)
        value = self.update_import_data(data, attr, value,
                                        attr_to_col, attr_to_main,
                                        self.ID_BY_NAME_COL,
                                        wb)
        self.update_md5(md5, value)

      md5 = md5.hexdigest()
      if is_new or orig_md5 != md5:
        for attr in single_value_attrs:
          if attr in related_data:
            related_data[attr] = related_data[attr][attr]

        model_name = data[self.ID_BY_NAME_COL].get(self.ID_BY_NAME_COL)
        if not model_name:
          continue

        model_data = name_map.get(model_name)
        if not model_data:
          continue

        model_data.setdefault(related_cs.related_attr, []).append(related_data)

  def _get_modified_data(self, obj, data, attr_to_col, is_new, is_update):
    """
    After data is read and processed from XLS cells, compare with existing
    object to get only changed values
    """
    modified = {}
    for attr, value in data.items():
      col = attr_to_col[attr]
      # import_val is current obj value formated for import, so that we can
      # compare with cell value
      import_val, current = col.data_for_import(obj)

      if not self.is_importable(obj, col, current):
        # for now we can't handle this type of values
        continue

      has_data_for_new = (is_new and (attr in data) and value)
      has_data_update = (
        not is_new
        and (value != import_val)
        and (value or import_val)) # last test ensures (None, u'', 0,
                                        #False, ...) are treated as equal
                                        #
                                        # FIXME: exception for bool columns to
                                        # check None != False?

      if is_update and not value:
        # update mode: we don't expect full records, and we don't erase
        # values with empty ones
        has_data_update = False

      if has_data_for_new or has_data_update:
        update = None
        try:
          update = self._import_value(obj, col, current, value)
        except ExcelImportError, e:
          update = col.UpdateCls(attr, current, value, None)
          update.error = True
          update.error_msg = e.message
          update.update = e.imported_value
        except ValueError, e:
          update = col.UpdateCls(attr, current, value, None)
          update.error = True
          update.error_msg = e.message

        if col.required:
          if update is None:
            # 'if update' may evaluate to False, beware to test existence with
            # 'is None'
            update = col.UpdateCls(attr, current, value, None)

          if update.value is None:
            if is_update:
              # update mode: we don't expect full records, and we don't erase
              # values with empty ones
              continue
            modified['__required_missing__'] = True
            if not update.error:
              # don't mask import error with 'required' message
              update.error = True
              update.error_msg = _(u'This field is required')

        if update:
          modified[attr] = update

    return modified


  def save_data(self, data):
    """
    @param data: list of (id(int), attr_sig, modified(dict))
    """
    session = db.session()
    changed_items = 0
    skipped_items = 0
    created_items = 0
    error_happened = False
    ids = [update.id for update in data if update.id is not None]
    q = self.model_cls.query

    if ids:
      q.filter(self.model_cls.id.in_(ids)).all()

    attr_to_column = self.attr_to_column

    for item_update in data:
      is_new = item_update.id is None
      item = q.get(item_update.id) if not is_new else self.model_cls()
      signed_attrs = '{}.{}'.format(u';'.join(item_update.attrs),
                                    item_update.sig)
      if not self.signer.validate(signed_attrs):
        item_id = str(item_update.id) if not is_new else 'new'
        logger.debug('Validation failed, skipping item "%s"'
                     '\n' 'item_attrs="%s"'
                     '\n' 'attr_sig="%s"',
                     item_id, repr(item_update.attrs), repr(item_update.sig))
        skipped_items += 1
        continue

      try:
        with session.begin_nested():
          for attr, value in item_update.data.items():
            if attr not in item_update.attrs:
              continue
            col = attr_to_column[attr]
            if hasattr(col, 'type_') and col.type_ is not None:
              value = col.type_(value)

            value = col.deserialize(value)
            import_val, current = col.data_for_import(item)
            update = self._import_value(item, col, current, value)
            # at this stage we don't expect anymore import errors
            assert not update.error

            if col.required and not update.value:
              # FIXME: missing param
              raise ExcelImportError(u'missing required')

            if update:
              self._set_obj_value(item, attr, update)

          db.session.add(item)

          # import "many relateds" values
          many_relateds = item_update.data.get('__many_related__', {})
          many_relateds_map = { cs.related_attr: cs for cs in self.MANY_RELATED_CS }

          for rel_attr, updates in many_relateds.items():
            cs = many_relateds_map.get(rel_attr)
            if not cs:
              logger.error('Many relateds: columns set for "%s" not found',
                           rel_attr,
                           extra={'stack': True,})
              continue

            manager = cs.create_manager()
            prop = sa.inspect(self.model_cls).attrs[cs.related_attr]
            # FIXME: we assume relation is made with 1 attribute
            prop = list(prop._reverse_property)[0]
            prop_key = prop.key
            del prop
            rel_attr_to_col = self.get_attr_to_column(cs, map_related_attr=True)

            for update in updates:
              obj = cs.model_cls()
              setattr(obj, prop_key, item)

              for attr, value in update.data.items():
                if attr not in update.attrs:
                  continue
                col = rel_attr_to_col[attr]
                if hasattr(col, 'type_') and col.type_ is not None:
                  value = col.type_(value)

                value = col.deserialize(value)
                import_val, current = col.data_for_import(obj)
                imported = manager._import_value(obj, col, current, value)
                # at this stage we don't expect anymore import errors
                assert not imported.error

                if col.required and not imported.value:
                  # FIXME: missing param
                  raise ExcelImportError(u'missing required')

                if imported:
                  manager._set_obj_value(obj, attr, imported)

              db.session.add(obj)

      except Exception as e:
        if isinstance(e, sa.exc.StatementError):
          logger.error('Import error: %s%s\n%s',
                       e.message, e.statement, pprint.pformat(e.params),
                       exc_info=True)
        else:
          logger.error('Import error: %s', unicode(e).encode('utf-8'),
                       exc_info=True)
          raise
        error_happened = True
        skipped_items += 1
        # skip this item
        continue
      else:
        if is_new:
          created_items += 1
        else:
          changed_items +=1

    if error_happened:
      #FIXME: include data for showing failed signatures, error during attr
      #conversion, etc
      logger.error('Excel import error', extra={'stack': True,})

    db.session.commit()
    return dict(changed_items=changed_items,
                created_items=created_items,
                skipped_items=skipped_items,
                error_happened=error_happened)

  def update_md5(self, md5, value):
    """ Compute consistent md5 for exported and imported value.

    For example if model value is None, exported value may be u''.
    All values are also converted to utf-8 encoded string before hashing.
    """
    if value is None or isinstance(value, list):
      value = u''

    value = unicode(value).encode('utf-8')
    md5.update(value)

  def extract_signature(self, signed, signer=None):
    """ Return the signature part of a signed string
    """
    if signer is None:
      signer = self.signer
    signature = signer.sep.join(signed.rsplit(signer.sep, 2)[-2:])
    logger.debug('extract_signature: "%s" from "%s"', repr(signature), signed)
    return signature

  def columns_for(self, field, form):
    if field.name in self.SKIP_COLS:
      return None

    custom_columns = 'columns_set_{}'.format(field.name)
    if hasattr(self, custom_columns):
      return getattr(self, custom_columns)(field, form)

    if isinstance(field, ModelFieldList):
      return None

    if hasattr(self.model_cls, field.name):
      attr = field.name
      label = field.label.text
      db_col = self.db_columns.get(field.name)
      column_cls, type_ = self.column_type(attr, db_col)
      required = u'required' in field.flags
      return column_cls(attr, label, type_, required)

    return None

  def column_type(self, attr, db_col):
    column_cls = None
    type_ = None
    custom_type = 'column_type_{}'.format(attr)
    if hasattr(self, custom_type):
      column_cls, type_ = getattr(self, custom_type)(attr, db_col)

    if column_cls is None:
      column_cls = Column

    if db_col is None:
      # relationship
      mapper = self.mapper
      prop = mapper.relationships.get(attr)

      if prop is not None and prop.direction is sa.orm.interfaces.MANYTOONE:
        target_class = prop.mapper.class_

        if issubclass(target_class, BaseVocabulary):
          column_cls = VocabularyColumn
          type_ = target_class.query.active().by_label
        elif issubclass(target_class, PostalAddress):
          column_cls = PostalAddressColumn
          type_ = None

    elif type_ is None and db_col is not None:
      try:
        type_ = db_col.type.python_type
        if type_ is bool:
          type_ = lambda v: v in (True, 'True', '1', 1)
        elif type_ in (datetime.date, datetime.datetime):
          # native excel type, no need to set cast function
          column_cls = (DateColumn
                        if type_ is datetime.date
                        else DateTimeColumn)
          type_ = None
      except NotImplementedError:
        pass

    return column_cls, type_


  def is_importable(self, item, column, current):
    if not column.importable:
      return False

    attr_name = (column.related_attr if isinstance(column, RelatedColumnSet)
                 else column.attr)

    custom_import = 'import_{}'.format(attr_name)
    if hasattr(self, custom_import):
      return True

    mapper = sa.inspect(item.__class__)
    relationship = mapper.relationships.get(attr_name)
    if relationship is not None:
      target_class = relationship.mapper.class_

      if (relationship.direction is sa.orm.interfaces.MANYTOONE
          and issubclass(target_class, BaseVocabulary)):
        return True

      return False # we cannot handle relationships in a generic way

    if isinstance(current, (db.Model, list)):
      return False

    try:
      col = getattr(self.db_columns, attr_name)
    except AttributeError:
      pass
    else:
      if isinstance(col.type, JSONType):
        return False

    return True

  def _import_value(self, obj, column, current, data):
    """
    Return a ::class `Update` instance.

    Lookup for method import_{attr}(obj, value). Expected returned value
    is (value_key, value). value_key should be data convertible to
    string (for web serialization) and should give the same value if
    _import_value is called with it as parameter 'data'

    If there's no method defined, value is returned as is.

    May raise ExcelImportError or ValueError with an explicit message

    For example given a string, update value could be en Entity
    """
    is_related = False
    attr_name = None

    if isinstance(data, Invalid):
      raise ExcelImportError(
        _(u'Invalid cell format, value: {value}').format(value=data.value),
        imported_value=data
        )

    if isinstance(column, RelatedColumnSet):
      is_related = True
      attr_name = column.related_attr
    else:
      attr_name = column.attr

    custom_import = 'import_{}'.format(attr_name)
    if hasattr(self, custom_import):
      data, value = getattr(self, custom_import)(obj, data)
      return column.UpdateCls(attr_name, current, data, value)

    value = data

    if is_related:
      raise ExcelImportError(
        _(u'Cannot import related entity for attribute "{}" '
          'without a custom importer method').format(attr_name),
        imported_value=value)

    try:
      col = getattr(self.db_columns, attr_name)
    except AttributeError:
      pass
    else:
      if 'choices' in col.info:
        # unicode(v) will make translatable values... translated
        rev_choices = {unicode(v): k for k,v in col.info.get('choices').items()}
        if value not in rev_choices:
          valid = u', '.join((u'"{}"'.format(v)
                              for v in rev_choices.keys() if v))
          raise ExcelImportError(
            _(u'"{value}" is invalid. Valid choices are: {valid}'
              ).format(value=value, valid=valid),
            imported_value=value,
          )
        value = rev_choices.get(value, value)

    return column.UpdateCls(attr_name, current, data, value)

  def _set_obj_value(self, obj, attr, update):
    """
    During save: set value in `update` on obj.

    This method is meant to be overriden for special attributes
    """
    setattr(obj, attr, update.value)

  def get_cell_value(self, sheet, row, col):
    """
    convert a cell value to useful python value returns a tuple of (value,
    excel cell type)
    """
    cell = sheet.cell(row=row, column=col)
    value = cell.value

    if isinstance(value, STRING_TYPES):
      value = cell.value.strip().replace(u'\r\n', u'\n')
      if u'\222' in value or u'\225' in value:
        # handle weird copy paste that may happen in excel: value is not actual
        # unicode but a cp1252 one that must be converted to unicode
        value = value.encode('raw_unicode_escape').decode('cp1252')
      elif value == '-':
        value = u""

    return value

  def update_import_data(self, data, attr, value, attr_map,
                         attr_to_main, ID_BY_NAME_COL, UNIQUE_ID_COLS, wb):
    """
    update data dict of imported value.
    (value, cell_types) comes from get_cell_value.

    @param wb: current workbook
    """
    column = attr_map[attr]

    if (value is not None
        and column.expected_cell_types is not None):
      if type(value) not in column.expected_cell_types:
        try:
          value = column.adapt_from_cell(value, wb)
        except:
          value = Invalid(value)

    if (not isinstance(value, Invalid)
        and value is not None and column.type_ is not None):
      try:
        value = column.type_(value)
      except:
        value = Invalid(value)

    main_col = attr_to_main[attr]
    main_col_name = (main_col.related_attr
                     if isinstance(main_col, RelatedColumnSet)
                     else column.attr)
    data[main_col_name][attr] = value

    if attr == ID_BY_NAME_COL and not isinstance(value, Invalid):
      data['__metadata__']['unique_name'] = value

    if attr in UNIQUE_ID_COLS:
      data['__metadata__']['unique_id_cols'][attr] = value

    return value
