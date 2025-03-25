# coding: utf-8
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.mixins import UserPassesTestMixin
from django.db.models import Q
from django.db import IntegrityError
from django.views import View
from django.views.generic import DetailView
from sortable_listview import SortableListView
from django.shortcuts import render, HttpResponseRedirect
from django.http import JsonResponse
from django.urls import reverse
from django.conf import settings
import datetime
from django.forms.models import model_to_dict
import posgradmin.forms as forms
from openpyxl import load_workbook
import posgradmin.models as models
from simple_search import search_filter
from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from pprint import pprint, pformat


class AcademicoSearch(View):
    def get(self, request, *args, **kwargs):
        qs = request.GET.get('qs', default=None)
        if qs is not None:
            search_fields = ['top_5',
                             'otras_publicaciones',
                             'lineas',
                             'palabras_clave',
                             'proyectos_sostenibilidad',
                             'proyectos_vigentes',
                             ]
            f = search_filter(search_fields, qs)
            results = models.Academico.objects.filter(f)

            response = JsonResponse({a.user.username: {'nombre': a.user.get_full_name(),
                                                    'palabras_clave': a.palabras_clave.split('\n')}
                                    for a in results})
            response["Access-Control-Allow-Origin"] = "*"
            return response
        else:
            response = JsonResponse({'aguas': 'wey'})
            response["Access-Control-Allow-Origin"] = "*"
            return response


class AcademicoInvitar(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        if self.request.user.is_superuser:
            return True

        elif self.request.user.is_staff:
            return True

    template_name = 'posgradmin/invita_candidatos.html'

    form_class = forms.AcademicoInvitarForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request,
                      self.template_name,
                      {'form': form,
                       'title': 'Crear candidatos',
                       })

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            wb = load_workbook(request.FILES['lista'])
            ws = wb.active

            i = 1
            rows = []
            while True:
                nombre = ws["A%s" % i].value
                apellidos = ws["B%s" % i].value
                email = ws["C%s" % i].value

                i += 1
                if (nombre is None
                        and apellidos is None
                        and email is None):
                    break
                elif (nombre == 'nombre'
                      and apellidos == 'apellidos'
                      and email == 'email'):
                    continue
                else:
                    rows.append([nombre, apellidos, email])

            errores = []
            aciertos = []
            i = 0
            for row in rows:
                i += 1
                [nombre, apellidos, email] = row
                try:
                    u = models.User()
                    u.first_name = nombre
                    u.last_name = apellidos
                    validate_email(email)
                    u.email = email
                    u.username = email.split('@')[0]
                    u.save()

                    a = models.Academico()
                    a.acreditacion = 'candidato'
                    a.user = u
                    a.save()
                    aciertos.append([i, ] + row)
                except (IntegrityError, ValidationError) as E:
                    errores.append([E.message, i] + [str(cell)
                                                     for cell in row])

            return render(request,
                          self.template_name,
                          {'form': form,
                           'title': 'Crear candidatos',
                           'form_errors': errores,
                           'aciertos': aciertos
                           })
        else:
            return render(request,
                          self.template_name,
                          {'form': form,
                           'title': 'Crear candidatos',
                           })


class InicioView(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'
    breadcrumbs = ((settings.APP_PREFIX + 'inicio/', 'Inicio'),)

    template_name = 'posgradmin/inicio.html'

    def get(self, request, *args, **kwargs):

        return render(request,
                      self.template_name,
                      {'title': 'Inicio',
                       'breadcrumbs': self.breadcrumbs})


class PerfilPublico(View):
    template = "posgradmin/perfil_publico.html"

    def get(self, request, *args, **kwargs):

        user = models.User.objects.get(username=kwargs['username'])

        breadcrumbs = ((settings.APP_PREFIX + 'inicio/', 'Inicio'),
                       (settings.APP_PREFIX + 'inicio/perfil/', 'Perfiles'),
                       (settings.APP_PREFIX
                        + 'inicio/perfil/%s' % user.get_username(),
                        user.get_full_name()))

        return render(request,
                      self.template,
                      {'U': user,
                       'title': user.get_full_name(),
                       'breadcrumbs': breadcrumbs})


class PerfilPublicoIndice(View):
    pass


class PerfilComite(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):

        if self.request.user.is_superuser:
            return True

        elif self.request.user.is_staff:
            return True

        elif hasattr(self.request.user, 'academico'):
            if self.request.user.academico.comite_academico:
                return True
            else:
                return False
        else:
            return False

    template = "posgradmin/perfil_comite.html"

    def get(self, request, *args, **kwargs):

        user = models.User.objects.get(username=kwargs['username'])

        user.academico.pc_resumen_academico()
        user.academico.wc_resumen_academico()
        a = user.academico
        a.semaforo_maestria = a.verifica_semaforo_maestria()
        a.semaforo_doctorado = a.verifica_semaforo_doctorado()
        a.save()


        breadcrumbs = ((settings.APP_PREFIX + 'inicio/', 'Inicio'),
                       (settings.APP_PREFIX + 'inicio/perfil/', 'Perfiles'),
                       (settings.APP_PREFIX
                        + 'inicio/perfil/%s' % user.get_username(),
                        user.get_full_name()))

        return render(request,
                      self.template,
                      {'U': user,
                       'title': user.get_full_name(),
                       'breadcrumbs': breadcrumbs})


class UserDetail(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    login_url = settings.APP_PREFIX + 'accounts/login/'
    model = models.User
    template_name = "posgradmin/perfil_personal.html"

    def test_func(self):
        if hasattr(self.request.user, 'asistente') \
           or hasattr(self.request.user, 'estudiante') \
           or hasattr(self.request.user, 'academico') \
           or self.request.user.is_staff:
            return True
        else:
            return False

    def get_context_data(self, **kwargs):
        context = super(UserDetail, self).get_context_data(**kwargs)
        context['user'] = self.request.user

        # may see intimacies?
        u = context['object']

        if (  # by authority
                self.request.user == u
                or self.request.user.is_staff
                or hasattr(self.request.user, 'asistente')):
            see_private = True
        else:
            see_private = False

        context['see_private'] = see_private

        return context


class PerfilDetail(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    login_url = settings.APP_PREFIX + 'accounts/login/'
    template_name = "posgradmin/perfil_personal.html"

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super(PerfilDetail, self).get_context_data(**kwargs)
        context['user'] = self.request.user

        # may see intimacies?
        u = context['object']

        if (  # by authority
                self.request.user == u
                or self.request.user.is_staff
                or hasattr(self.request.user, 'asistente')):
            see_private = True
        else:
            see_private = False

        context['see_private'] = see_private

        return context

    def get_object(self):
        return self.request.user


class PerfilEditar(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    form_class = forms.PerfilModelForm

    breadcrumbs = ((settings.APP_PREFIX + 'inicio/', 'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil/', 'Mi perfil'),
                   (settings.APP_PREFIX + 'inicio/perfil/editar/', 'Editar'))

    template = 'posgradmin/perfil_editar.html'

    def get(self, request, *args, **kwargs):
        try:
            perfil = request.user.perfil
            data = model_to_dict(perfil)
            data['fecha_nacimiento'] = data['fecha_nacimiento'].isoformat()
            data['nombre'] = request.user.first_name
            data['apellidos'] = request.user.last_name
            form = self.form_class(data=data)
        except:
            data = {'nombre': request.user.first_name,
                    'apellidos': request.user.last_name}
            form = self.form_class(data=data)
            return render(request,
                          self.template,
                          {'form': form,
                           'title': 'Perfil Personal',
                           'breadcrumbs': self.breadcrumbs})

        return render(request,
                      self.template,
                      {'object': perfil,
                       'form': form,
                       'title': 'Editar Perfil Personal',
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            u = request.user
            u.first_name = request.POST['nombre']
            u.last_name = request.POST['apellidos']
            u.save()
            try:
                p = request.user.perfil
            except:
                p = models.Perfil()

            p.user = request.user

            fecha_nacimiento = datetime.date(
                int(request.POST['fecha_nacimiento_year']),
                int(request.POST['fecha_nacimiento_month']),
                int(request.POST['fecha_nacimiento_day']))
            p.fecha_nacimiento = fecha_nacimiento
            p.genero = request.POST['genero']
            p.nacionalidad = request.POST['nacionalidad']
            p.curp = request.POST['curp']
            p.rfc = request.POST['rfc']
            p.telefono = request.POST['telefono']
            p.telefono_movil = request.POST['telefono_movil']
            p.website = request.POST['website']
            p.direccion1 = request.POST['direccion1']
            p.codigo_postal = request.POST['codigo_postal']
            if 'headshot' in request.FILES:
                p.headshot = request.FILES['headshot']
            elif 'headshot-clear' in request.POST and a.headshot.name != '':
                a.headshot.delete()
            p.save()

            return HttpResponseRedirect(reverse('perfil'))
        else:
            return render(request,
                          self.template,
                          {'object': u.perfil,
                           'form': form,
                           'title': 'Editar mi perfil',
                           'breadcrumbs': self.breadcrumbs})


class PerfilAcademicoDetail(LoginRequiredMixin, UserPassesTestMixin,
                            DetailView):
    login_url = settings.APP_PREFIX + 'accounts/login/'
    template_name = "posgradmin/perfil_academico.html"

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super(PerfilAcademicoDetail, self).get_context_data(**kwargs)
        context['user'] = self.request.user

        # may see intimacies?
        u = context['object']

        if (  # by authority
                self.request.user == u
                or self.request.user.is_staff):
            see_private = True
        else:
            see_private = False

        context['see_private'] = see_private

        return context

    def get_object(self):
        return self.request.user




class PerfilEstudianteDetail(LoginRequiredMixin, UserPassesTestMixin,
                             DetailView):
    login_url = settings.APP_PREFIX + 'accounts/login/'
    template_name = "posgradmin/perfil_estudiante.html"

    def test_func(self):

        return True

    def get_context_data(self, **kwargs):
        context = super(PerfilEstudianteDetail, self).get_context_data(**kwargs)
        context['user'] = self.request.user

        # may see intimacies?
        u = context['object']

        if (  # by authority
                self.request.user == u  # 'tis mine
                or self.request.user.is_staff):  # we ait'
            see_private = True
        else:
            see_private = False

        context['see_private'] = see_private

        return context

    def get_object(self):
        return self.request.user


class AcademicoPerfilView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    form_class = forms.AcademicoPerfilModelForm

    breadcrumbs = ((settings.APP_PREFIX + 'inicio/',
                    'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil-academico/',
                    'Perfil Académico'),
                   (settings.APP_PREFIX + 'inicio/academico/perfil',
                    'Editar perfil general'))

    template_name = 'posgradmin/try.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        if hasattr(request.user, 'academico'):
            a = request.user.academico
            data = model_to_dict(a)
            form = self.form_class(data=data, instance=a)
        else:
            form = self.form_class()

        return render(request,
                      self.template_name,
                      {'form': form,
                       'title': 'Editar Perfil Académico',
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():

            a = request.user.academico

            a.CVU = request.POST[u'CVU']
            a.nivel_SNI = request.POST[u'nivel_SNI']
            a.estimulo_UNAM = request.POST[u'estimulo_UNAM']

            if 'anexo_CV' in request.FILES:
                a.anexo_CV = request.FILES['anexo_CV']
            elif 'anexo_CV-clear' in request.POST and a.anexo_CV.name != '':
                a.anexo_CV.delete()

            if 'anexo_solicitud' in request.FILES:
                a.anexo_solicitud = request.FILES['anexo_solicitud']
            elif 'anexo_solicitud-clear' in request.POST and a.anexo_solicitud.name != '':
                a.anexo_solicitud.delete()

            if 'ultimo_grado' in request.FILES:
                a.ultimo_grado = request.FILES['ultimo_grado']
            elif 'ultimo_grado-clear' in request.POST and a.ultimo_grado.name != '':
                a.ultimo_grado.delete()


            # if request.POST['participacion_tutor_doctorado'] != "":
            #     a.participacion_tutor_doctorado = request.POST['participacion_tutor_doctorado']
            # if request.POST['participacion_comite_doctorado'] != "":
            #     a.participacion_comite_doctorado = request.POST['participacion_comite_doctorado']

            # if request.POST['participacion_tutor_maestria'] != "":
            #     a.participacion_tutor_maestria = request.POST['participacion_tutor_maestria']
            # if request.POST['participacion_comite_maestria'] != "":
            #     a.participacion_comite_maestria = request.POST['participacion_comite_maestria']

            if 'disponible_miembro' in request.POST:
                a.disponible_miembro = True
            else:
                a.disponible_miembro = False

            if 'disponible_tutor' in request.POST:
                a.disponible_tutor = True
            else:
                a.disponible_tutor = False




            a.save()

            return HttpResponseRedirect(reverse('perfil_academico'))
        else:
            return render(request,
                          self.template_name,
                          {'form': form,
                           'title': 'Editar Perfil Académico',
                           'breadcrumbs': self.breadcrumbs})


class AcademicoResumenCVView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    form_class = forms.AcademicoResumenCVModelForm

    breadcrumbs = ((settings.APP_PREFIX + 'inicio/',
                    'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil-academico/',
                    'Perfil Académico'),
                   (settings.APP_PREFIX + 'inicio/academico/resumen',
                    'Editar resumen curricular'))

    template_name = 'posgradmin/try.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        if hasattr(request.user, 'academico'):
            a = request.user.academico
            data = model_to_dict(a)
            form = self.form_class(data=data, instance=a)
        else:
            form = self.form_class()

        return render(request,
                      self.template_name,
                      {'form': form,
                       'title': 'Editar resumen curricular',
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():

            a = request.user.academico

            if request.POST['tesis_doctorado'] != "":
                a.tesis_doctorado = int(request.POST['tesis_doctorado'])
            if request.POST['tesis_doctorado_5'] != "":
                a.tesis_doctorado_5 = int(request.POST['tesis_doctorado_5'])

            if request.POST['tesis_maestria'] != "":
                a.tesis_maestria = int(request.POST['tesis_maestria'])
            if request.POST['tesis_maestria_5'] != "":
                a.tesis_maestria_5 = int(request.POST['tesis_maestria_5'])

            if request.POST['tesis_licenciatura'] != "":
                a.tesis_licenciatura = int(request.POST['tesis_licenciatura'])
            if request.POST['tesis_licenciatura_5'] != "":
                a.tesis_licenciatura_5 = int(request.POST['tesis_licenciatura_5'])

            if request.POST['tutor_principal_otros_programas']:
                a.tutor_principal_otros_programas = request.POST['tutor_principal_otros_programas']

            if request.POST['comite_doctorado_otros'] != "":
                a.comite_doctorado_otros = int(request.POST['comite_doctorado_otros'])
            if request.POST['comite_maestria_otros'] != "":
                a.comite_maestria_otros = int(request.POST['comite_maestria_otros'])

            a.otras_actividades = request.POST['otras_actividades']

            if request.POST['articulos_internacionales_5'] != "":
                a.articulos_internacionales_5 = int(request.POST['articulos_internacionales_5'])
            if request.POST['articulos_nacionales_5'] != "":
                a.articulos_nacionales_5 = int(request.POST['articulos_nacionales_5'])
            if request.POST['articulos_internacionales'] != "":
                a.articulos_internacionales = int(request.POST['articulos_internacionales'])
            if request.POST['articulos_nacionales'] != "":
                a.articulos_nacionales = int(request.POST['articulos_nacionales'])
            if request.POST['capitulos'] != "":
                a.capitulos = int(request.POST['capitulos'])
            if request.POST['capitulos_5'] != "":
                a.capitulos_5 = int(request.POST['capitulos_5'])
            if request.POST['libros'] != "":
                a.libros = int(request.POST['libros'])
            if request.POST['libros_5'] != "":
                a.libros_5 = int(request.POST['libros_5'])
            a.top_5 = request.POST['top_5']

            a.otras_publicaciones = request.POST['otras_publicaciones']

            a.semaforo_maestria = a.verifica_semaforo_maestria()
            a.semaforo_doctorado = a.verifica_semaforo_doctorado()


            print a.semaforo_maestria, a.semaforo_doctorado

            a.save()

            print a.publicaciones_5()
            return HttpResponseRedirect(reverse('perfil_academico'))
        else:
            return render(request,
                          self.template_name,
                          {'form': form,
                           'title': 'Editar Resumen Curricular',
                           'breadcrumbs': self.breadcrumbs})



class AcademicoActividadView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    form_class = forms.AcademicoActividadModelForm

    breadcrumbs = ((settings.APP_PREFIX + 'inicio/',
                    'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil-academico/',
                    'Perfil Académico'),
                   (settings.APP_PREFIX + 'inicio/academico/actividad',
                    'Editar actividad profesional y de investigación'))

    template_name = 'posgradmin/try.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        if hasattr(request.user, 'academico'):
            a = request.user.academico
            data = model_to_dict(a)
            form = self.form_class(data=data, instance=a)
        else:
            form = self.form_class()

        return render(request,
                      self.template_name,
                      {'form': form,
                       'title': 'Editar Actividad Profesional',
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)

        if form.is_valid():

            a = request.user.academico

            a.lineas = request.POST['lineas']

            if 'lineas_de_investigacion' in request.POST:
                lineas = []
                for l in request.POST.getlist('lineas_de_investigacion'):
                    lid = int(l)
                    linea = models.LineaInvestigacion.objects.get(pk=lid)
                    lineas.append(linea)
                a.lineas_de_investigacion.set(lineas)
            else:
                a.lineas_de_investigacion.set([])

            if 'campos_de_conocimiento' in request.POST:
                campos = []
                for c in request.POST.getlist('campos_de_conocimiento'):
                    cid = int(c)
                    campo = models.CampoConocimiento.objects.get(pk=cid)
                    campos.append(campo)
                a.campos_de_conocimiento.set(campos)
            else:
                a.campos_de_conocimiento.set([])

            a.palabras_clave = request.POST['palabras_clave']
            a.motivacion = request.POST['motivacion']
            a.proyectos_sostenibilidad = request.POST['proyectos_sostenibilidad']
            a.proyectos_vigentes = request.POST['proyectos_vigentes']
            if 'disponible_miembro' in request.POST:
                a.disponible_miembro = True
            else:
                a.disponible_miembro = False

            if 'disponible_tutor' in request.POST:
                a.disponible_tutor = True
            else:
                a.disponible_tutor = False


            a.semaforo_maestria = a.verifica_semaforo_maestria()
            a.semaforo_doctorado = a.verifica_semaforo_doctorado()

            a.save()

            return HttpResponseRedirect(reverse('perfil_academico'))
        else:
            return render(request,
                          self.template_name,
                          {'form': form,
                           'title': 'Editar Perfil Académico',
                           'breadcrumbs': self.breadcrumbs})


class GradoAcademicoAgregar(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    form_class = forms.GradoAcademicoModelForm

    breadcrumbs = [(settings.APP_PREFIX + 'inicio/', 'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil/', 'Mi perfil'),
                   (settings.APP_PREFIX
                    + 'inicio/perfil/agregar-grado',
                    'Agregar Grado Académico')]

    template_name = 'posgradmin/grado_academico_agregar.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request,
                      self.template_name,
                      {'title': 'Agregar Grado Académico',
                       'form': form,
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            ins = models.Institucion.objects.get(
                id=int(request.POST['institucion']))
            g = models.GradoAcademico(
                user=request.user,
                nivel=request.POST['nivel'],
                grado_obtenido=request.POST['grado_obtenido'],
                institucion=ins,
                fecha_obtencion=datetime.date(
                    int(request.POST['fecha_obtencion_year']),
                    int(request.POST['fecha_obtencion_month']),
                    int(request.POST['fecha_obtencion_day'])))
            g.save()

            return HttpResponseRedirect(reverse('perfil'))
        else:
            return render(request,
                          self.template_name,
                          {'title': 'Agregar Grado Académico',
                           'form': form,
                           'breadcrumbs': self.breadcrumbs})


class GradoAcademicoEliminar(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def get(self, request, *args, **kwargs):
        g = models.GradoAcademico.objects.get(id=int(kwargs['pk']))

        if request.user == g.user:
            g.delete()

        return HttpResponseRedirect(reverse('perfil'))


class AdscripcionAgregar(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    form_class = forms.AdscripcionModelForm

    breadcrumbs = [(settings.APP_PREFIX + 'inicio/', 'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil/', 'Perfil Personal'),
                   (settings.APP_PREFIX + 'inicio/perfil/agregar-adscrpcion',
                    'Agregar Adscripción')]

    template_name = 'posgradmin/try.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request,
                      self.template_name,
                      {'title': 'Agregar Adscripcion',
                       'form': form,
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            ins = models.Institucion.objects.get(
                id=int(request.POST['institucion']))
            if 'catedra_conacyt' in request.POST:
                catedra_conacyt = True
            else:
                catedra_conacyt = False

            if hasattr(request.user, 'perfil'):
                a = models.Adscripcion(
                    perfil=request.user.perfil,
                    nombramiento=request.POST['nombramiento'],
                    anno_nombramiento=request.POST['anno_nombramiento'],
                    catedra_conacyt=catedra_conacyt,
                    institucion=ins)
                a.save()
            else:
                return HttpResponseRedirect(reverse('editar_perfil'))

            if ins.entidad_PCS or request.user.perfil.asociado_PCS():
                return HttpResponseRedirect(reverse('perfil'))
            else:
                return HttpResponseRedirect(reverse('agregar_asociacion'))

        else:
            return render(request,
                          self.template_name,
                          {'title': 'Agregar Adscripción',
                           'form': form,
                           'breadcrumbs': self.breadcrumbs})


class AsociacionAgregar(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    form_class = forms.AsociacionModelForm

    breadcrumbs = [(settings.APP_PREFIX + 'inicio/', 'Inicio'),
                   (settings.APP_PREFIX + 'inicio/perfil/', 'Perfil Personal'),
                   (settings.APP_PREFIX + 'inicio/perfil/agregar-asociacion',
                    'Agregar Asociación al PCS')]

    template_name = 'posgradmin/try.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request,
                      self.template_name,
                      {'title': 'Agregar Asociación al PCS',
                       'form': form,
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            ins = models.Institucion.objects.get(
                id=int(request.POST['institucion']))

            assert ins.entidad_PCS

            a = models.Adscripcion(
                perfil=request.user.perfil,
                institucion=ins,
                nombramiento='asociación',
                anno_nombramiento=datetime.datetime.now().year,
                asociacion_PCS=True)
            a.save()
            return HttpResponseRedirect(reverse('perfil'))

        else:
            print "aga"
            return render(request,
                          self.template_name,
                          {'title': 'Agregar Asociación al PCS',
                           'form': form,
                           'breadcrumbs': self.breadcrumbs})


class AdscripcionEliminar(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def get(self, request, *args, **kwargs):
        a = models.Adscripcion.objects.get(id=int(kwargs['pk']))
        if request.user == a.perfil.user:
            a.delete()
        return HttpResponseRedirect(reverse('perfil'))


class InstitucionAgregarView(LoginRequiredMixin, View):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    form_class = forms.InstitucionModelForm

    breadcrumbs = [(settings.APP_PREFIX + 'inicio/', 'Inicio'),
                   (settings.APP_PREFIX +
                    'institucion/agregar', 'Agregar Institución')]

    template_name = 'posgradmin/institucion_agregar.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request,
                      self.template_name,
                      {'title': 'Agregar Institución',
                       'form': form,
                       'breadcrumbs': self.breadcrumbs})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            i = models.Institucion(
                nombre=request.POST['nombre'],
                pais=request.POST['pais'],
                estado=request.POST['estado'],
                suborganizacion=request.POST['suborganizacion'],
                dependencia_UNAM=True
                if 'dependencia_UNAM' in request.POST else False,
                entidad_PCS=False)
            i.save()
            if kwargs['devolver'] == 'ga':
                return HttpResponseRedirect(reverse('agregar_grado'))
            elif kwargs['devolver'] == 'ad':
                return HttpResponseRedirect(reverse('agregar_adscripcion'))
        else:
            return render(request,
                          self.template_name,
                          {'title': 'Agregar Institución',
                           'form': form,
                           'breadcrumbs': self.breadcrumbs})


class EstudianteSortableView(LoginRequiredMixin,
                             UserPassesTestMixin,
                             SortableListView):

    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    def get_queryset(self):

        print self.request.encoding
        self.request.GET.encoding = 'utf-8'
        search = self.request.GET.get('search', None)
        qs = super(EstudianteSortableView, self).get_queryset()

        if search:
            search = search.encode('utf-8')
            sorted = qs.filter(Q(cuenta__icontains=search) |
                               Q(user__first_name__icontains=search) |
                               Q(user__last_name__icontains=search))
        else:
            sorted = qs
        return sorted

    allowed_sort_fields = {'user': {'default_direction': '',
                                    'verbose_name': 'nombre'},
                           'estado': {'default_direction': '-',
                                      'verbose_name': 'estado'},
                           'ingreso': {'default_direction': '-',
                                       'verbose_name': 'año de ingreso'}}
    default_sort_field = 'user'

    paginate_by = 15

    model = models.Estudiante


class AcademicoSortableView(LoginRequiredMixin,
                            UserPassesTestMixin,
                            SortableListView):
    login_url = settings.APP_PREFIX + 'accounts/login/'

    def test_func(self):
        return True

    # def get_queryset(self):
    #     sorted = super(EstudianteSortableView, self).get_queryset()

    allowed_sort_fields = {'user': {'default_direction': '',
                                    'verbose_name': 'nombre'},
                           'acreditacion': {'default_direction': '-',
                                            'verbose_name':
                                            'acreditación'}}
    default_sort_field = 'user'

    paginate_by = 15

    model = models.Academico


# class CatedraSortableView(LoginRequiredMixin,
#                           UserPassesTestMixin,
#                           SortableListView):
#     login_url = settings.APP_PREFIX + 'accounts/login/'

#     def test_func(self):
#         return True

#     def get_queryset(self):
#         sorted = super(CatedraSortableView, self).get_queryset()
#         return sorted

#     allowed_sort_fields = {'profesor': {'default_direction': '',
#                                         'verbose_name': 'profesor'},
#                            'curso': {'default_direction': '-',
#                                      'verbose_name': 'curso'},
#                            'semestre': {'default_direction': '-',
#                                         'verbose_name': 'semestre'},
#                            'year': {'default_direction': '-',
#                                     'verbose_name': 'año'}}
#     default_sort_field = 'curso'

#     paginate_by = 15

#     model = models.Catedra
