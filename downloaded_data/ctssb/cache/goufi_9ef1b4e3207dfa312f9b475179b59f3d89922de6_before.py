# -*- coding: utf-8 -*-
'''
Created on 17 may 2018

@author: C. Guychard
@copyright: ©2018 Article 714
@license: AGPL v3
'''

"""
a set of classes to be used in mixins for processor that provide support for importing XLS* files
"""

from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from xlrd import open_workbook
from xlrd.book import Book


from odoo.addons.goufi_base.utils.converters import toString

from .processor import MultiSheetLineIterator


#-------------------------------------------------------------------------------------
# CONSTANTS
XL_AUTHORIZED_EXTS = ('xlsx', 'xls')


class XLImporterBaseProcessor(MultiSheetLineIterator):

    def __init__(self, parent_config):
        super(XLImporterBaseProcessor, self).__init__(parent_config)
        self.book = None
    #-------------------------------------------------------------------------------------

    def _open_xls(self, import_file):
        # returns Book
        return open_workbook(import_file.filename)

    #-------------------------------------------------------------------------------------

    def _open_xlsx(self, import_file):
        # returns Workbook
        return load_workbook(import_file.filename, read_only=True, keep_vba=False, guess_types=False, data_only=True)

    #-------------------------------------------------------------------------------------
    def get_book(self, import_file):
        """
        Method that actually process data
        """

        self.logger.info(" process XLS* file: " + toString(import_file.filename))
        try:

            if import_file.filename.endswith('.xls'):
                result = self._open_xls(import_file)
            elif import_file.filename.endswith('.xlsx'):
                # retu
                result = self._open_xlsx(import_file)
            return result

        except Exception as e:
            self.logger.exception("Not able to open XL file: %s", str(e))
            self.odooenv.cr.rollback()
            self.errorCount += 1
            return False

    #-------------------------------------------------------------------------------------
    # tab generator
    def get_tabs(self, import_file=None):
        if isinstance(self.book, Book):
            for shname in self.book.sheet_names():
                yield (shname, self.book.sheet_by_name(shname))
        elif isinstance(self.book, Workbook):
            for shname in self.book.sheetnames:
                yield (shname, self.book[shname])
        else:
            self.logger.error("Unrecognized Book type....")

    #-------------------------------------------------------------------------------------
    # line generator
    def get_rows(self, tab=None):
        if isinstance(self.book, Book):
            for index in range(tab[1].nrows):
                # Filter empty values
                notempty = True
                rv = tab[1].row_values(index)
                for v in rv:
                    notempty = notempty or (v != None and v != '')
                if notempty:
                    yield (index, tab[1].row(index))
        elif isinstance(self.book, Workbook):
            # Empty cells ou value None
            index = 0
            for row in tab[1]:
                notempty = False
                for c in row:
                    v = c.value
                    notempty = notempty or not isinstance(c, EmptyCell) or (v != None and v != '')
                if notempty:
                    yield (index, row)
                index += 1
        else:
            self.logger.error("Unrecognized Book type....")

    #-------------------------------------------------------------------------------------
    # Process header for tab
    def process_tab_header(self, tab=None, headerrow=None):

        if isinstance(self.book, Book):
            header = tab[1].row_values(headerrow[0])
            return header
        elif isinstance(self.book, Workbook):
            header = []
            for c in headerrow[1]:
                if c.value != None:
                    header.append(c.value)
                else:
                    # stop whenever an empty cell is found in header
                    break
            return header
        else:
            self.logger.error("Unrecognized Book type....")
            return None

    #-------------------------------------------------------------------------------------
    # Provides a dictionary of values in a row
    def get_row_values(self, tab=None, row=None):

        if isinstance(self.book, Book):
            row_vals = tab[1].row_values(row[0])
            return row_vals
        elif isinstance(self.book, Workbook):
            values = []
            for c in row[1]:
                values.append(c.value)
            return values
        else:
            self.logger.error("Unrecognized Book type....")
            return None

    #-------------------------------------------------------------------------------------
    # Provides a dictionary of values in a row
    def get_row_values_as_dict(self, tab=None, row=None, tabheader=None):

        if isinstance(self.book, Book):
            values = {}
            row_vals = tab[1].row_values(row[0])
            hsize = len(tabheader)
            for idx in range(0, hsize):
                values[tabheader[idx]] = row_vals[idx]
            return values
        elif isinstance(self.book, Workbook):
            values = {}
            for c in row[1]:
                colname = None
                if not isinstance(c, EmptyCell) and not c.column == None:
                    colname = tabheader[c.column - 1]
                if colname != None:
                    values[colname] = c.value
            return values
        else:
            self.logger.error("Unrecognized Book type....")
            return None

    #-------------------------------------------------------------------------------------
    def process_file(self, import_file, force=False):
        ext = import_file.filename.split('.')[-1]
        if (ext in XL_AUTHORIZED_EXTS):
            super(XLImporterBaseProcessor, self).process_file(import_file, force)
        else:
            self.logger.error("Cannot process file: Wrong extension -> %s", ext)

    #-------------------------------------------------------------------------------------
    def process_data(self, import_file):
        """
        Method that actually process data
        """
        self.book = self.get_book(import_file)

        if self.book != None:
            super(XLImporterBaseProcessor, self).process_data(import_file)
        if isinstance(self.book, Book):
            self.book.release_resources()
        self.book = None
