from Acquisition import aq_inner
from Acquisition import aq_parent
from collective.transmogrifier.transmogrifier import Transmogrifier
from opengever.base.interfaces import IReferenceNumberPrefix
from opengever.bundle.console import add_guid_index
from opengever.bundle.ldap import DisabledLDAP
from opengever.bundle.sections.bundlesource import BUNDLE_PATH_KEY
from opengever.bundle.sections.commit import INTERMEDIATE_COMMITS_KEY
from opengever.dossier.behaviors.dossier import IDossierMarker
from opengever.maintenance.debughelpers import setup_app
from opengever.maintenance.debughelpers import setup_plone
from opengever.maintenance.scripts.update_object_ids import ObjectIDUpdater
from opengever.repository.behaviors import referenceprefix
from opengever.repository.interfaces import IRepositoryFolder
from opengever.repository.interfaces import IRepositoryFolderRecords
from opengever.setup.sections.xlssource import xlrd_xls2array
from openpyxl import Workbook
from openpyxl.styles import Font
from plone import api
from plone.app.uuid.utils import uuidToCatalogBrain
from plone.app.uuid.utils import uuidToObject
from plone.uuid.interfaces import IUUID
from Products.CMFPlone.utils import safe_unicode
from uuid import uuid4
from zope.annotation import IAnnotations
import argparse
import json
import shutil
import sys
import tempfile


class OperationItem(object):

    def __init__(self, position=None, title=None, description=None):
        self.position = self.cleanup_position(position)
        self.title = self.to_safe_unicode(title)
        self.description = self.to_safe_unicode(description)

    @staticmethod
    def to_safe_unicode(maybe_none):
        if maybe_none is None:
            return None
        maybe_none = safe_unicode(maybe_none)
        if maybe_none:
            return maybe_none

    @staticmethod
    def cleanup_position(position):
        """Remove splitting dots - they're not usefull for comparison.
        This only works for grouped_by_three formatter.
        """
        if position is None:
            return None
        position = str(position)
        if position:
            return position.replace('.', '')

    @property
    def reference_number_prefix(self):
        """Returns last part of the position - the referencenumber prefix"""
        return self.position[-1]

    @property
    def parent_position(self):
        """Returns the position without the last part of the position, i.e.
        the parent position"""
        return self.position[:-1]

    def __repr__(self):
        return u"OperationItem({}, {}, {})".format(self.position, self.title, self.description).encode('utf-8')

    def __eq__(self, other):
        return all((self.position == other.position,
                    self.title == other.title,
                    self.description == other.description))


class Row(object):

    def __init__(self, row, column_mapping):
        for key, column in column_mapping.items():
            col = column['index']
            setattr(self, key, row[col])


class ExcelDataExtractor(object):

    header_row = 2
    first_data_row = 6
    column_mapping = {
        'old_position': {'index': 0, 'header': u'Ordnungs-\npositions-\nnummer'},
        'old_title': {'index': 1, 'header': u'Titel der Ordnungsposition'},
        'old_description': {'index': 2, 'header': u'Beschreibung (optional)'},
        'new_position': {'index': 5, 'header': u'Ordnungs-\npositions-\nnummer'},
        'new_title': {'index': 6, 'header': u'Titel der Ordnungsposition'},
        'new_description': {'index': 8, 'header': u'Beschreibung (optional)'},
    }

    def __init__(self, diff_xlsx_path):
        sheets = xlrd_xls2array(diff_xlsx_path)
        self.data = sheets[0]['sheet_data']
        self.validate_format()

    def validate_format(self):
        headers = self.data[self.header_row]
        for column in self.column_mapping.values():
            assert headers[column['index']] == column['header'], \
                u"Column header mismatch: {} != {}".format(
                    headers[column['index']], column['header'])

    def get_data(self):
        for row in self.data[self.first_data_row:]:
            yield Row(row, self.column_mapping)


class RepositoryExcelAnalyser(object):

    def __init__(self, mapping_path, analyse_path):
        self.number_changes = {}

        self.diff_xlsx_path = mapping_path
        self.analyse_xlsx_path = analyse_path
        self.analysed_rows = []
        self._reference_repository_mapping = None
        self.catalog = api.portal.get_tool('portal_catalog')

        # A mapping new_position_number:UID
        self.position_uid_mapping = {}

        # A mapping new_position_number:guid
        self.position_guid_mapping = {}

    def analyse(self):
        data_extractor = ExcelDataExtractor(self.diff_xlsx_path)

        for row in data_extractor.get_data():
            new_item = {}
            if row.new_position in ['', u'l\xf6schen', '-']:
                # Position should be deleted
                new_item = OperationItem()
            else:
                new_item = OperationItem(row.new_position, row.new_title, row.new_description)
            if row.old_position == '':
                # Position did not exist
                old_item = OperationItem()
            else:
                old_item = OperationItem(row.old_position, row.old_title, row.old_description)

            # Ignore empty rows
            if not old_item.position and not new_item.position:
                continue

            new_number = None
            new_parent_position = None
            new_parent_uid = None

            parent_of_new_position = None
            new_position_guid = None

            needs_creation = not bool(old_item.position)
            need_number_change, need_move = self.needs_number_change_or_move(new_item, old_item)
            if need_number_change:
                new_number = self.get_new_number(new_item)
            if need_move:
                new_parent_position, new_parent_uid = self.get_new_parent_position_and_uid(new_item)
            if needs_creation:
                parent_of_new_position = self.get_parent_of_new_position(new_item)
                new_position_guid = uuid4().hex[:8]

            analyse = {
                'uid': self.get_uuid_for_position(old_item.position),
                'new_position': parent_of_new_position,
                'new_position_guid': new_position_guid,
                'new_title': self.get_new_title(new_item, old_item) if not needs_creation else None,
                'new_number': new_number,
                'new_parent_position': new_parent_position,
                'new_parent_uid': new_parent_uid,
                'old_item': old_item,
                'new_item': new_item,
                'repository_depth_violated': self.is_repository_depth_violated(
                    new_item, old_item),
                'leaf_node_violated': need_move and self.is_leaf_node_principle_violated(
                    new_item, old_item)
            }

            self.analysed_rows.append(analyse)
            if not needs_creation:
                self.position_uid_mapping[new_item.position] = analyse['uid']
            else:
                self.position_guid_mapping[new_item.position] = new_position_guid

    def get_new_title(self, new_item, old_item):
        """Returns the new title or none if no rename is necessary."""
        if new_item.title != old_item.title:
            return new_item.title

        return None

    def get_new_number(self, new_item):
        """Returns latest part of the position - the new referencenumber
        prefix"""
        return new_item.reference_number_prefix

    def get_new_parent_position_and_uid(self, new_item):
        """Returns the new parent position and the uid. If the object does not
        yet exists it returns the guid."""

        parent_position = new_item.parent_position
        if parent_position not in self.position_uid_mapping:
            # Parent does not exist yet and will be created in the
            # first step of the migration
            return parent_position, self.position_guid_mapping[parent_position]

        return parent_position, self.position_uid_mapping[parent_position]

    def get_parent_of_new_position(self, new_item):
        final_parent_position = new_item.parent_position

        parent_row = [item for item in self.analysed_rows
                      if item['new_item'].position == final_parent_position]

        if not parent_row:
            return final_parent_position

        # The parent will be moved to the right position so we need to add
        # the subrepofolder on the "old position"
        return parent_row[0]['old_item'].position

    def needs_move(self, new_item, old_item):
        if not old_item.position or not new_item.position:
            return False

        if new_item.parent_position != old_item.parent_position:
            return True

        return False

    def needs_number_change_or_move(self, new_item, old_item):
        """Check if a number change or even a move is necessary
        """
        need_number_change = False
        need_move = False

        if new_item.position and old_item.position:
            if new_item.position != old_item.position:
                need_number_change = True
                self.number_changes[new_item.position] = old_item.position

                # check if parent is already changed - so no need to change
                parent_position = new_item.parent_position
                if parent_position in self.number_changes:
                    if self.number_changes[parent_position] == old_item.parent_position:
                        need_number_change = False

                if need_number_change:
                    # check if move is necessary
                    if new_item.parent_position != old_item.parent_position:
                        need_move = True

        return need_number_change, need_move

    def is_repository_depth_violated(self, new_item, old_item):
        max_depth = api.portal.get_registry_record(
            interface=IRepositoryFolderRecords, name='maximum_repository_depth')

        if new_item.position and len(new_item.position) > max_depth:
            return True

        return False

    def is_leaf_node_principle_violated(self, new_item, old_item):
        parent_number = new_item.parent_position
        parent_repo = self.get_repository_reference_mapping().get(parent_number)
        if not parent_repo:
            # Parent does not exist yet, so nothing to worry about it
            return False

        has_dossiers = any([IDossierMarker.providedBy(item) for item in
                            parent_repo.objectValues()])
        return has_dossiers

    def get_repository_reference_mapping(self):
        if not self._reference_repository_mapping:
            repos = [brain.getObject() for brain in
                     self.catalog(object_provides=IRepositoryFolder.__identifier__)]
            self._reference_repository_mapping = {
                repo.get_repository_number(): repo for repo in repos}

        return self._reference_repository_mapping

    def get_uuid_for_position(self, position):
        mapping = self.get_repository_reference_mapping()

        if position and position in mapping:
            return IUUID(mapping[position])

        return None

    def export_to_excel(self):
        workbook = self.prepare_workbook(self.analysed_rows)
        # Save the Workbook-data in to a StringIO
        return workbook.save(filename=self.analyse_xlsx_path)

    def prepare_workbook(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Analyse'

        self.insert_label_row(sheet)
        self.insert_value_rows(sheet, rows)

        return workbook

    def insert_label_row(self, sheet):
        title_font = Font(bold=True)
        labels = [
            # metadata
            'Neu: Position', 'Neu: Titel', 'Neu: Description',
            'Alt: Position', 'Alt: Titel', 'Alt: Description',

            # operations
            'Position Erstellen (Parent Aktenzeichen)',
            'Umbenennung (Neuer Titel)',
            'Nummer Anpassung (Neuer `Praefix`)',
            'Verschiebung (Aktenzeichen neues Parent)',

            # rule violations
            'Verletzt Max. Tiefe',
            'Verletzt Leafnode Prinzip'
        ]

        for i, label in enumerate(labels, 1):
            cell = sheet.cell(row=1 + 1, column=i)
            cell.value = label
            cell.font = title_font

    def insert_value_rows(self, sheet, rows):
        for row, data in enumerate(rows, 2):
            values = [
                data['new_item'].position,
                data['new_item'].title,
                data['new_item'].description,
                data['old_item'].position,
                data['old_item'].title,
                data['old_item'].description,
                data['new_position'],
                data['new_title'],
                data['new_number'],
                data['new_parent_position'],
                'x' if data['repository_depth_violated'] else '',
                'x' if data['leaf_node_violated'] else '',
            ]

            for column, attr in enumerate(values, 1):
                cell = sheet.cell(row=1 + row, column=column)
                cell.value = attr


class RepositoryMigrator(object):

    def __init__(self, operations_list):
        self.operations_list = operations_list
        self._reference_repository_mapping = None
        self.catalog = api.portal.get_tool('portal_catalog')

    def run(self):
        self.create_repository_folders(self.items_to_create())
        self.move_branches(self.items_to_move())
        self.adjust_reference_number_prefix(self.items_to_adjust_number())
        self.rename(self.items_to_rename())
        self.update_description(self.operations_list)
        self.reindex()
        # self.validate()

    def items_to_create(self):
        return [item for item in self.operations_list if item['new_position']]

    def items_to_move(self):
        return [item for item in self.operations_list if item['new_parent_position']]

    def items_to_adjust_number(self):
        return [item for item in self.operations_list if item['new_number']]

    def items_to_rename(self):
        return [item for item in self.operations_list if item['new_title']]

    def create_repository_folders(self, items):
        """Add repository folders - by using the ogg.bundle import. """

        bundle_items = []
        for item in items:
            # Bundle expect the format [[repository], [dossier]]
            parent_reference = [[int(x) for x in list(item['new_position'])]]
            bundle_items.append(
                {'guid': item['new_position_guid'],
                 'description': item['new_item'].description,
                 'parent_reference': parent_reference,
                 'reference_number_prefix': item['new_item'].reference_number_prefix,
                 'review_state': 'repositoryfolder-state-active',
                 'title_de': item['new_item'].title})

        tmpdirname = tempfile.mkdtemp()
        with open('{}/repofolders.json'.format(tmpdirname), 'w') as _file:
            json.dump(bundle_items, _file)

        self.start_bundle_import(tmpdirname)

        shutil.rmtree(tmpdirname)

    def start_bundle_import(self, bundle_path):
        add_guid_index()

        portal = api.portal.get()
        transmogrifier = Transmogrifier(portal)
        ann = IAnnotations(transmogrifier)
        ann[BUNDLE_PATH_KEY] = bundle_path
        ann[INTERMEDIATE_COMMITS_KEY] = False

        with DisabledLDAP(portal):
            transmogrifier(u'opengever.bundle.oggbundle')

    def move_branches(self, items):
        mapping = self.get_repository_reference_mapping()
        for item in items:
            parent = uuidToObject(item['new_parent_uid'])
            if not parent:
                parent = self.catalog(guid=item['new_parent_uid'])[0].getObject()

            repo = uuidToObject(item['uid'])
            if not parent or not repo:
                raise Exception('No parent or repo found for {}'.format(item))

            api.content.move(source=repo, target=parent, safe_id=True)

    def get_repository_reference_mapping(self):
        if not self._reference_repository_mapping:
            repos = [brain.getObject() for brain in
                     self.catalog(object_provides=IRepositoryFolder.__identifier__)]
            self._reference_repository_mapping = {
                repo.get_repository_number(): repo for repo in repos}

        return self._reference_repository_mapping

    def adjust_reference_number_prefix(self, items):
        parents = set()
        for item in items:
            repo = uuidToObject(item['uid'])
            referenceprefix.IReferenceNumberPrefix(repo).reference_number_prefix = item['new_number']
            parents.add(aq_parent(aq_inner(repo)))

        self.regenerate_reference_number_mapping(list(parents))

    def regenerate_reference_number_mapping(self, objs):
        for obj in objs:
            ref_adapter = IReferenceNumberPrefix(obj)
            # This purges also the dossier mapping, but the parents does not
            # contain any dossier otherwise something is wrong and an
            # exception will be raised when looping over the childs.
            ref_adapter.purge_mappings()

            for child in obj.listFolderContents():
                if not IRepositoryFolder.providedBy(child):
                    raise Exception(
                        'A parent of a repositoryfolder contains dossiers')
                ref_adapter.set_number(
                    child, number=child.reference_number_prefix)

    def rename(self, items):
        for item in items:
            repo = uuidToObject(item['uid'])

            # Rename
            repo.title_de = item['new_title']

            # Adjust id if necessary
            ObjectIDUpdater(repo, FakeOptions()).maybe_update_id()

    def update_description(self, items):
        for item in items:
            repo = uuidToObject(item['uid'])
            if not repo:
                continue

            new_description = item['new_item'].description
            if repo.description != new_description:
                repo.description = new_description

    def reindex(self):
        for item in self.operations_list:
            obj = uuidToObject(item['uid'])
            if not obj:
                # New created objects can be ignored
                break

            obj.reindexObject(idxs=['Title', 'sortable_title', 'path',
                                    'reference', 'Description'])

    def validate(self):
        """This steps make sure that the repository system has
        been correctly migrated."""

        for item in self.operations_list:
            obj = uuidToObject(item['uid'])

            # Assert reference number, title and description on the object
            assert item['new_item'].position == obj.get_repository_number()
            assert item['new_item'].title == obj.title_de
            assert item['new_item'].description == obj.description

            # Assert catalog
            brain = uuidToCatalogBrain(item['uid'])
            expected_title = u'{} {}'.format(
                item['new_item'].position, item['new_item'].title)
            assert expected_title == brain.title_de


class FakeOptions(object):
    dry_run = False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', dest='site_root', default=None,
                        help='Absolute path to the Plone site')
    parser.add_argument('-m', dest='mapping', default=None,
                        help='Path to the mapping xlsx')
    parser.add_argument('-o', dest='output', default=None,
                        help='Path to the output xlsx')
    options = parser.parse_args(sys.argv[3:])
    app = setup_app()

    setup_plone(app, options)

    analyser = RepositoryExcelAnalyser(options.mapping, options.output)
    analyser.analyse()
    analyser.export_to_excel()


if __name__ == '__main__':
    main()
