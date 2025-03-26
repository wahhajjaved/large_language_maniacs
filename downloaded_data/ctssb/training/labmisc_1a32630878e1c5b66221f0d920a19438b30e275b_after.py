#!/usr/bin/env python3

import sys

import numpy as np
import openpyxl as xl
import pandas as pd

ROW_DATA_STARTS = 41
ROW_ANALYTE_NAME = 39

def main(filename):
    workbook = xl.load_workbook(filename)
    ws = workbook.active
    invisible_rows = [key for key in
            sorted(ws.row_dimensions.keys()) if
            not ws.row_dimensions[key].visible]
    data_columns = np.where([cell.value for cell in ws.rows[39]])[0] + 1
    data_rows = np.where([cell.value for cell in ws.columns[1]])[0] + 1
    data_rows = [row for row in data_rows if row >= ROW_DATA_STARTS and row not in invisible_rows]
    
    col_names = data_columns[0]
    col_dilution = data_columns[1]

    col_last = max(data_columns)
    col = data_columns[2]
    data = []
    while col <= col_last:
        analyte = ws.cell(row=ROW_ANALYTE_NAME, column=col).value
        for row in data_rows:
            datum = {"Analyte": analyte, "Error": None}
            datum["Sample"] = ws.cell(row=row, column=col_names).value
            datum["Dilution"] = ws.cell(row=row, column=col_dilution).value
            intensity_cell = ws.cell(row=row, column=col)
            datum["Intensity"] = intensity_cell.value if intensity_cell.data_type == "n" else None
            conc_cell = ws.cell(row=row, column=col+1)
            if conc_cell.data_type == "n":
                datum["Concentration"] = conc_cell.value
                if conc_cell.style.font.color.theme != 1:
                    datum["Error"] = "OOR"
            else:
                datum["Error"] = conc_cell.value
            datum["Expected"] = ws.cell(row=row, column=col+2).value
            data.append(datum)
        col += 3

    pd.DataFrame(data).to_csv(sys.stdout, index=False)


if __name__ == "__main__":
    main(sys.argv[1])
