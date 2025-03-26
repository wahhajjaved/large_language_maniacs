from Acquisition import aq_inner
from Acquisition import aq_parent
from collections import defaultdict
from collections import namedtuple
from collective.transmogrifier.transmogrifier import Transmogrifier
from opengever.base.indexes import sortable_title
from opengever.base.interfaces import IReferenceNumber
from opengever.base.interfaces import IReferenceNumberFormatter
from opengever.base.interfaces import IReferenceNumberPrefix
from opengever.bundle.console import add_guid_index
from opengever.bundle.ldap import DisabledLDAP
from opengever.bundle.sections.bundlesource import BUNDLE_PATH_KEY
from opengever.bundle.sections.commit import INTERMEDIATE_COMMITS_KEY
from opengever.bundle.sections.constructor import BUNDLE_GUID_KEY
from opengever.dossier.behaviors.dossier import IDossierMarker
from opengever.globalindex.handlers.task import TaskSqlSyncer
from opengever.maintenance.debughelpers import setup_app
from opengever.maintenance.debughelpers import setup_plone
from opengever.maintenance.scripts.update_object_ids import ObjectIDUpdater
from opengever.repository.behaviors import referenceprefix
from opengever.repository.deleter import RepositoryDeleter
from opengever.repository.interfaces import IRepositoryFolder
from opengever.repository.interfaces import IRepositoryFolderRecords
from opengever.setup.sections.xlssource import xlrd_xls2array
from openpyxl import Workbook
from openpyxl.styles import Font
from plone import api
from plone.app.uuid.utils import uuidToCatalogBrain
from plone.app.uuid.utils import uuidToObject
from plone.uuid.interfaces import IUUID
from Products.CMFPlone.utils import safe_unicode
from uuid import uuid4
from zope.annotation import IAnnotations
from zope.component import queryAdapter
import argparse
import json
import logging
import shutil
import sys
import tempfile


logger = logging.getLogger('migration')
logger.setLevel(logging.INFO)


class MigrationPreconditionsError(Exception):
    """Raised when errors are found during migration validation"""


class MigrationValidationError(Exception):
    """Raised when errors are found during migration validation"""


class OperationItem(object):

    def __init__(self, position=None, title=None, description=None):
        self.position = self.cleanup_position(position)
        self.title = self.to_safe_unicode(title)
        self.description = self.to_safe_unicode(description)

    @staticmethod
    def to_safe_unicode(maybe_none):
        if maybe_none is None:
            return None
        maybe_none = safe_unicode(maybe_none)
        if maybe_none:
            return maybe_none

    @staticmethod
    def cleanup_position(position):
        """Remove splitting dots - they're not usefull for comparison.
        This only works for grouped_by_three formatter.
        """
        if position is None:
            return None
        position = str(position)
        if position:
            return position.replace('.', '')

    @property
    def reference_number_prefix(self):
        """Returns last part of the position - the referencenumber prefix"""
        return self.position[-1]

    @property
    def parent_position(self):
        """Returns the position without the last part of the position, i.e.
        the parent position"""
        return self.position[:-1]

    def __repr__(self):
        return u"OperationItem({}, {}, {})".format(self.position, self.title, self.description).encode('utf-8')

    def __eq__(self, other):
        return all((self.position == other.position,
                    self.title == other.title,
                    self.description == other.description))


class Row(object):

    def __init__(self, row, column_mapping):
        for key, column in column_mapping.items():
            col = column.index
            setattr(self, key, row[col])


Column = namedtuple('Column', ('index', 'technical_header', 'header'))


class ExcelDataExtractor(object):

    header_row = 2
    technical_header_row = 4
    first_data_row = 6

    column_mapping = {
        'old_position': Column(0, '', u'Ordnungs-\npositions-\nnummer'),
        'old_title': Column(1, '', u'Titel der Ordnungsposition'),
        'old_description': Column(2, '', u'Beschreibung (optional)'),
        'new_position': Column(5, 'reference_number', u'Ordnungs-\npositions-\nnummer'),
        'new_title': Column(6, 'effective_title', u'Titel der Ordnungsposition'),
        'new_description': Column(8, 'description', u'Beschreibung (optional)'),
        'block_inheritance': Column(19, 'block_inheritance', ''),
        'read': Column(20, 'read_dossiers_access', ''),
        'add': Column(21, 'add_dossiers_access', ''),
        'edit': Column(22, 'edit_dossiers_access', ''),
        'close': Column(23, 'close_dossiers_access', ''),
        'reactivate': Column(24, 'reactivate_dossiers_access', ''),
        'manage_dossiers': Column(25, 'manage_dossiers_access', ''),
    }

    def __init__(self, diff_xlsx_path):
        sheets = xlrd_xls2array(diff_xlsx_path)
        self.data = sheets[0]['sheet_data']
        self.validate_format()

    def validate_format(self):
        headers = self.data[self.header_row]
        technical_headers = self.data[self.technical_header_row]
        for column in self.column_mapping.values():
            assert technical_headers[column.index] == column.technical_header, \
                u"Column technical header mismatch: {} != {}".format(
                    technical_headers[column.index], column.technical_header)

            if not column.header:
                continue

            assert headers[column.index] == column.header, \
                u"Column header mismatch: {} != {}".format(
                    headers[column.index], column.header)

    def get_data(self):
        for row in self.data[self.first_data_row:]:
            yield Row(row, self.column_mapping)


class RepositoryExcelAnalyser(object):

    def __init__(self, mapping_path, analyse_path):
        self.number_changes = {}

        self.diff_xlsx_path = mapping_path
        self.analyse_xlsx_path = analyse_path
        self.analysed_rows = []
        self._reference_repository_mapping = None
        self.final_positions = []
        self.catalog = api.portal.get_tool('portal_catalog')

        # A mapping new_position_number:UID
        self.position_uid_mapping = {}

        # A mapping new_position_number:guid
        self.position_guid_mapping = {}

        self.check_preconditions()
        self.reporoot, self.reporoot_guid = self.prepare_guids()

    def check_preconditions(self):
        # current implementation only works with grouped_by_three reference
        # formatter, notably because we remove splitting dots during the analysis.
        formatter_name = api.portal.get_registry_record(
            "opengever.base.interfaces.IReferenceNumberSettings.formatter")
        assert formatter_name == "grouped_by_three", "Migration is only supported with grouped_by_three"
        self.formatter = queryAdapter(
                api.portal.get(), IReferenceNumberFormatter, name=formatter_name)

        # Creation of new repository folders in the repository root will only
        # work if their is a single repository root.
        results = self.catalog.unrestrictedSearchResults(
            portal_type='opengever.repository.repositoryroot')
        assert len(results) == 1, "Migration is only supported with a single repository root"

    def prepare_guids(self):
        """ The GUID index is needed by the bundle transmogrifier.
        Moreover the repository root needs to have a GUID, as it does not
        have a reference number allowing to find the parent when creating
        a new repository folder in the repository root.
        """
        add_guid_index()
        brain = self.catalog.unrestrictedSearchResults(
            portal_type='opengever.repository.repositoryroot')[0]
        reporoot = brain.getObject()

        if not IAnnotations(reporoot).get(BUNDLE_GUID_KEY):
            IAnnotations(reporoot)[BUNDLE_GUID_KEY] = uuid4().hex[:8]
        reporoot.reindexObject(idxs=['bundle_guid'])
        return reporoot, IAnnotations(reporoot).get(BUNDLE_GUID_KEY)

    def analyse(self):
        data_extractor = ExcelDataExtractor(self.diff_xlsx_path)

        for row in data_extractor.get_data():
            new_item = {}
            if row.new_position in ['', u'l\xf6schen', '-']:
                # Position should be deleted
                new_item = OperationItem()
            else:
                new_item = OperationItem(row.new_position, row.new_title, row.new_description)
            if row.old_position == '':
                # Position did not exist
                old_item = OperationItem()
            else:
                old_item = OperationItem(row.old_position, row.old_title, row.old_description)

            # Ignore empty rows
            if not old_item.position and not new_item.position:
                continue

            # Skip positions that should be deleted
            if not new_item.position:
                logger.info("Skipping, we do not support deletion", row)
                continue

            new_number = None
            new_parent_position = None
            new_parent_uid = None
            merge_into = None

            new_position_parent_position = None
            new_position_parent_guid = None
            new_position_guid = None

            permissions = None

            needs_creation = not bool(old_item.position)
            need_number_change, need_move, need_merge = self.needs_number_change_move_or_merge(new_item, old_item)

            if need_number_change:
                new_number = self.get_new_number(new_item)
            if need_move:
                new_parent_position, new_parent_uid = self.get_new_parent_position_and_uid(new_item)
            if need_merge:
                merge_into = self.get_position_and_uid_to_merge_into(new_item)
            if needs_creation:
                new_position_parent_position, new_position_parent_guid = self.get_parent_of_new_position(new_item)
                new_position_guid = uuid4().hex[:8]
                permissions = self.extract_permissions(row)

            operation = {
                'uid': self.get_uuid_for_position(old_item.position),
                'new_position_parent_position': new_position_parent_position,
                'new_position_parent_guid': new_position_parent_guid,
                'new_position_guid': new_position_guid,
                'new_title': self.get_new_title(new_item, old_item) if not needs_creation else None,
                'new_number': new_number,
                'new_parent_position': new_parent_position,
                'new_parent_uid': new_parent_uid,
                'merge_into': merge_into,
                'old_item': old_item,
                'new_item': new_item,
                'permissions': permissions
            }
            self.validate_operation(operation)

            self.analysed_rows.append(operation)
            if need_merge:
                pass
            elif not needs_creation:
                self.position_uid_mapping[new_item.position] = operation['uid']
            else:
                self.position_guid_mapping[new_item.position] = new_position_guid

    def validate_operation(self, operation):
        """Make sure that operation satisfies all necessary conditions and add
        is_valid, repository_depth_violated and leaf_node_violated to it.
        """
        operation['is_valid'] = True

        # Each operation should either have a uid or a new_position_guid
        if not any((operation['new_position_guid'], operation['uid'])):
            logger.warning("Invalid operation: needs new_position_guid "
                           "or uid. {}".format(operation))
            operation['is_valid'] = False

        # Each operation should have new position
        if not operation['new_item'].position:
            logger.warning("Invalid operation: needs new position. {}".format(
                operation))
            operation['is_valid'] = False

        if all((operation['new_position_guid'], operation['uid'])):
            logger.warning("Invalid operation: can define only one of "
                           "new_position_guid or uid. {}".format(operation))
            operation['is_valid'] = False

        # A move operation should have a new_parent_uid
        if operation['new_parent_position'] or operation['new_parent_uid']:
            if not operation['new_parent_uid']:
                logger.warning(
                    "Invalid operation: move operation must define "
                    "new_parent_uid. {}".format(operation))
                operation['is_valid'] = False

        # Make sure that if a position is being created, its parent will be found
        if not bool(operation['old_item'].position) and not operation['new_position_parent_guid']:
            parent = self.get_object_for_position(
                operation['new_position_parent_position'])

            if not parent:
                logger.warning(
                    "Invalid operation: could not find new parent for create "
                    "operation. {}".format(operation))
                operation['is_valid'] = False

        self.check_repository_depth_violation(operation)
        self.check_leaf_node_principle_violation(operation)

    def get_new_title(self, new_item, old_item):
        """Returns the new title or none if no rename is necessary."""
        if new_item.title != old_item.title:
            return new_item.title

        return None

    def get_new_number(self, new_item):
        """Returns latest part of the position - the new referencenumber
        prefix"""
        return new_item.reference_number_prefix

    def get_new_parent_position_and_uid(self, new_item):
        """Returns the new parent position and the uid. If the object does not
        yet exists it returns the guid."""

        parent_position = new_item.parent_position
        if not parent_position:
            # We are moving into the reporoot
            return parent_position, self.reporoot.UID()

        if parent_position not in self.position_uid_mapping:
            # Parent does not exist yet and will be created in the
            # first step of the migration
            return parent_position, self.position_guid_mapping.get(parent_position)

        return parent_position, self.position_uid_mapping[parent_position]

    def get_position_and_uid_to_merge_into(self, new_item):
        """Returns the position and the uid this should be merged into.
        If the object does not yet exists it returns the guid."""

        position = new_item.position

        if position not in self.position_uid_mapping:
            # Parent does not exist yet and will be created in the
            # first step of the migration
            return self.position_guid_mapping.get(position)

        return self.position_uid_mapping[position]

    def get_parent_of_new_position(self, new_item):
        final_parent_position = new_item.parent_position
        if not final_parent_position:
            # We are creating a new position in the reporoot
            return final_parent_position, self.reporoot_guid

        parent_row = [item for item in self.analysed_rows
                      if item['new_item'].position == final_parent_position]

        if not parent_row:
            # bundle import (ConstructorSection) will find parent from
            # the reference number
            return final_parent_position, None

        # Two possibilities, the new parent is being created or moved.
        if parent_row[0]['old_item'].position:
            # The parent will be moved to the right position so we need to add
            # the subrepofolder on the "old position"
            return parent_row[0]['old_item'].position, None
        else:
            # The parent is being created, so we will identify it through its guid.
            return None, parent_row[0]['new_position_guid']

    def needs_number_change_move_or_merge(self, new_item, old_item):
        """Check if a number change, a move or a merge is necessary
        """
        need_number_change = False
        need_move = False
        need_merge = False

        if new_item.position and old_item.position and new_item.position != old_item.position:
            self.number_changes[new_item.position] = old_item.position

            # If the new position is already in position_uid_mapping or
            # position_guid_mapping, it means that a previous operation will
            # move or create a repository_folder to that position. This means
            # this operation is a merge into an existing position.
            if (new_item.position in self.position_uid_mapping or
                    new_item.position in self.position_guid_mapping):
                need_merge = True
                return need_number_change, need_move, need_merge

            # check if move is necessary
            new_parent = new_item.parent_position
            old_parent = old_item.parent_position
            if new_parent != old_parent:
                need_move = True
                # check whether the parent is being moved
                if new_parent in self.number_changes:
                    if self.number_changes[new_parent] == old_parent:
                        need_move = False

            # a position that really needs to get moved will need its reference
            # number reindexed
            if need_move:
                need_number_change = True

            # check if number change is necessary
            if new_item.reference_number_prefix != old_item.reference_number_prefix:
                need_number_change = True

        return need_number_change, need_move, need_merge

    def check_repository_depth_violation(self, operation):
        max_depth = api.portal.get_registry_record(
            interface=IRepositoryFolderRecords, name='maximum_repository_depth')

        new_item = operation['new_item']
        if new_item.position and len(new_item.position) > max_depth:
            logger.warning(
                "Invalid operation: repository depth violated. {}".format(operation))
            operation['is_valid'] = False
            operation['repository_depth_violated'] = True
        else:
            operation['repository_depth_violated'] = False

    def check_leaf_node_principle_violation(self, operation):
        operation['leaf_node_violated'] = False
        if not (operation['new_parent_uid'] or operation['new_position_guid']):
            # object is neither moved nor created, nothing to worry about
            return

        parent_number = operation['new_item'].parent_position
        parent_repo = self.get_object_for_position(parent_number)
        if parent_repo and any([IDossierMarker.providedBy(item) for item in
                                parent_repo.objectValues()]):
            operation['is_valid'] = False
            operation['leaf_node_violated'] = True
            logger.warning("Invalid operation: leaf node principle violated. {}".format(operation))

    def get_repository_reference_mapping(self):
        if not self._reference_repository_mapping:
            repos = [brain.getObject() for brain in
                     self.catalog(object_provides=IRepositoryFolder.__identifier__)]
            self._reference_repository_mapping = {
                repo.get_repository_number().replace('.', ''): repo for repo in repos}

        return self._reference_repository_mapping

    def get_uuid_for_position(self, position):
        mapping = self.get_repository_reference_mapping()

        if position and position in mapping:
            return IUUID(mapping[position])

        return None

    def get_object_for_position(self, position):
        mapping = self.get_repository_reference_mapping()
        return mapping.get(position)

    def extract_permissions(self, row):
        permissions = {'block_inheritance': False}

        if row.block_inheritance:
            block = row.block_inheritance.strip()
            assert block in ['ja', 'nein']
            if block == 'ja':
                permissions['block_inheritance'] = True

        for key in ['read', 'add', 'edit', 'close', 'reactivate', 'manage_dossiers']:
            groups = [group.strip() for group in getattr(row, key).split(',')]
            groups = [group for group in groups if group]

            permissions[key] = groups

        return permissions

    def export_to_excel(self):
        workbook = self.prepare_workbook(self.analysed_rows)
        # Save the Workbook-data in to a StringIO
        return workbook.save(filename=self.analyse_xlsx_path)

    def prepare_workbook(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Analyse'

        self.insert_label_row(sheet)
        self.insert_value_rows(sheet, rows)

        return workbook

    def insert_label_row(self, sheet):
        title_font = Font(bold=True)
        labels = [
            # metadata
            'Neu: Position', 'Neu: Titel', 'Neu: Description',
            'Alt: Position', 'Alt: Titel', 'Alt: Description',

            # operations
            'Position Erstellen (Parent Aktenzeichen oder GUID)',
            'Umbenennung (Neuer Titel)',
            'Nummer Anpassung (Neuer `Praefix`)',
            'Verschiebung (Aktenzeichen neues Parent)',
            'Merge mit (UID oder GUID)',

            # rule violations
            'Verletzt Max. Tiefe',
            'Verletzt Leafnode Prinzip',
            'Ist ungultig',

            # permission
            'Bewilligungen',
        ]

        for i, label in enumerate(labels, 1):
            cell = sheet.cell(row=1 + 1, column=i)
            cell.value = label
            cell.font = title_font

    def insert_value_rows(self, sheet, rows):
        for row, data in enumerate(rows, 2):
            values = [
                data['new_item'].position,
                data['new_item'].title,
                data['new_item'].description,
                data['old_item'].position,
                data['old_item'].title,
                data['old_item'].description,
                data['new_position_parent_position'] or data['new_position_parent_guid'],
                data['new_title'],
                data['new_number'],
                data['new_parent_position'],
                data['merge_into'],
                'x' if data['repository_depth_violated'] else '',
                'x' if data['leaf_node_violated'] else '',
                'x' if not data['is_valid'] else '',
                data['permissions'],
            ]

            for column, attr in enumerate(values, 1):
                cell = sheet.cell(row=1 + row, column=column)
                cell.value = attr


class RepositoryMigrator(object):

    def __init__(self, operations_list):
        self.operations_list = operations_list
        self._reference_repository_mapping = None
        self.to_reindex = defaultdict(set)
        self.catalog = api.portal.get_tool('portal_catalog')
        self.check_preconditions()

    def check_preconditions(self):
        if any(not operation['is_valid'] for operation in self.operations_list):
            raise MigrationPreconditionsError("Some operations are invalid.")

    def run(self):
        self.create_repository_folders(self.items_to_create())
        self.move_branches(self.items_to_move())
        self.merge_branches(self.items_to_merge())
        self.adjust_reference_number_prefix(self.items_to_adjust_number())
        self.rename(self.items_to_rename())
        self.update_description(self.operations_list)
        self.reindex()
        self.validate()

    def items_to_create(self):
        return [item for item in self.operations_list if item['new_position_guid']]

    def items_to_move(self):
        return [item for item in self.operations_list if item['new_parent_uid']]

    def items_to_merge(self):
        return [item for item in self.operations_list if item['merge_into']]

    def items_to_adjust_number(self):
        return [item for item in self.operations_list if item['new_number']]

    def items_to_rename(self):
        return [item for item in self.operations_list if item['new_title']]

    def add_to_reindexing_queue(self, uid, idxs, with_children=False):
        self.to_reindex[uid].update(idxs)
        obj = uuidToObject(uid)
        if not with_children:
            return

        contained_brains = self.catalog.unrestrictedSearchResults(
            path=obj.absolute_url_path())
        for brain in contained_brains:
            self.to_reindex[brain.UID].update(idxs)

    def create_repository_folders(self, items):
        """Add repository folders - by using the ogg.bundle import. """
        bundle_items = []
        for item in items:
            # Bundle expect the format [[repository], [dossier]]
            parent_reference = None
            if item['new_position_parent_position']:
                parent_reference = [[int(x) for x in list(item['new_position_parent_position'])]]

            bundle_items.append(
                {'guid': item['new_position_guid'],
                 'description': item['new_item'].description,
                 'parent_reference': parent_reference,
                 'parent_guid': item['new_position_parent_guid'],
                 'reference_number_prefix': item['new_item'].reference_number_prefix,
                 'review_state': 'repositoryfolder-state-active',
                 'title_de': item['new_item'].title,
                 '_permissions': item['permissions']
                 })

        tmpdirname = tempfile.mkdtemp()
        with open('{}/repofolders.json'.format(tmpdirname), 'w') as _file:
            json.dump(bundle_items, _file)

        self.start_bundle_import(tmpdirname)

        shutil.rmtree(tmpdirname)

    def start_bundle_import(self, bundle_path):
        portal = api.portal.get()
        transmogrifier = Transmogrifier(portal)
        ann = IAnnotations(transmogrifier)
        ann[BUNDLE_PATH_KEY] = bundle_path
        ann[INTERMEDIATE_COMMITS_KEY] = False

        with DisabledLDAP(portal):
            transmogrifier(u'opengever.bundle.oggbundle')

    def uid_or_guid_to_object(self, uid_or_guid):
        obj = uuidToObject(uid_or_guid)
        if not obj:
            obj = self.catalog(bundle_guid=uid_or_guid)[0].getObject()
        return obj

    def move_branches(self, items):
        for item in items:
            parent = self.uid_or_guid_to_object(item['new_parent_uid'])
            repo = uuidToObject(item['uid'])
            if not parent or not repo:
                raise Exception('No parent or repo found for {}'.format(item))

            api.content.move(source=repo, target=parent, safe_id=True)

    def merge_branches(self, items):
        for item in items:
            target = self.uid_or_guid_to_object(item['merge_into'])
            repo = uuidToObject(item['uid'])
            if not target or not repo:
                raise Exception('No target or repo found for {}'.format(item))

            for obj in repo.contentValues():
                api.content.move(source=obj, target=target, safe_id=True)
                self.add_to_reindexing_queue(
                    obj.UID(), ('Title', 'sortable_title', 'reference'),
                    with_children=True)

            deleter = RepositoryDeleter(repo)
            if not deleter.is_deletion_allowed():
                raise Exception('Trying to delete not empty object {}'.format(item))
            deleter.delete()

    def adjust_reference_number_prefix(self, items):
        parents = set()
        for item in items:
            repo = uuidToObject(item['uid'])
            referenceprefix.IReferenceNumberPrefix(repo).reference_number_prefix = item['new_number']
            parents.add(aq_parent(aq_inner(repo)))
            self.add_to_reindexing_queue(
                item['uid'], ('Title', 'sortable_title', 'reference'),
                with_children=True)

        self.regenerate_reference_number_mapping(list(parents))

    def regenerate_reference_number_mapping(self, objs):
        for obj in objs:
            ref_adapter = IReferenceNumberPrefix(obj)
            # This purges also the dossier mapping, but the parents does not
            # contain any dossier otherwise something is wrong and an
            # exception will be raised when looping over the childs.
            ref_adapter.purge_mappings()

            for child in obj.listFolderContents():
                if not IRepositoryFolder.providedBy(child):
                    raise Exception(
                        'A parent of a repositoryfolder contains dossiers')
                ref_adapter.set_number(
                    child, number=child.reference_number_prefix)

    def rename(self, items):
        for item in items:
            repo = uuidToObject(item['uid'])

            # Rename
            repo.title_de = item['new_title']

            # Adjust id if necessary
            ObjectIDUpdater(repo, FakeOptions()).maybe_update_id()

            # We do not need to reindex path as this seems to already happen
            # recursively
            self.add_to_reindexing_queue(
                item['uid'], ('Title', 'sortable_title'))

    def update_description(self, items):
        for item in items:
            repo = uuidToObject(item['uid'])
            if not repo:
                continue

            new_description = item['new_item'].description
            if repo.description != new_description:
                repo.description = new_description
                self.add_to_reindexing_queue(item['uid'], ('Description',))

    def reindex(self):
        for uid, idxs in self.to_reindex.items():
            obj = uuidToObject(uid)
            obj.reindexObject(idxs=idxs)
            if obj.portal_type == 'opengever.task.task':
                # make sure that the model is up to date.
                TaskSqlSyncer(obj, None).sync()

    def guid_to_object(self, guid):
        results = self.catalog.unrestrictedSearchResults(bundle_guid=guid)
        if len(results) == 0:
            # This should never happen. Object with a guid should have been created.
            logger.warning(
                u"Couldn't find object with GUID %s in catalog" % guid)
            return

        if len(results) > 1:
            # Ambiguous GUID - this should never happen
            logger.warning(
                u"Ambiguous GUID! Found more than one result in catalog "
                u"for GUID %s " % guid)
            return

        return results[0].getObject()

    def validate(self):
        """This steps make sure that the repository system has
        been correctly migrated."""
        self.validation_errors = defaultdict(list)
        self.validation_failed = False

        for operation in self.operations_list:
            # Three possibilities here: position was created, deleted or modified
            if operation['new_position_guid']:
                # new position was created
                obj = self.guid_to_object(operation['new_position_guid'])
            elif operation['uid']:
                obj = uuidToObject(operation['uid'])
                if operation['merge_into']:
                    # position was deleted
                    if obj:
                        logger.error(u"Positions wasn't deleted correctly {}.".format(operation['uid']))
                        self.validation_failed = True
                    continue
            else:
                logger.error(u"Invalid operation {}".format(operation))
                self.validation_failed = True
                continue

            if not obj:
                uid = operation['new_position_guid'] or operation['uid']
                logger.error(u"Could not resolve object {}. Skipping validation.".format(uid))
                self.validation_failed = True
                continue

            # Assert reference number, title and description on the object
            uid = obj.UID()
            new = operation['new_item']
            self.assertEqual(uid, new.position, obj.get_repository_number(), 'incorrect number')
            self.assertEqual(uid, new.title, obj.title_de, 'incorrect title')
            self.assertEqual(uid, new.description, obj.description, 'incorrect description')

            # Assert that data in the catalog is consistent with data on the object
            self.assertObjectConsistency(obj)

        if self.validation_failed:
            raise MigrationValidationError("See log for details")

    def assertObjectConsistency(self, obj):
        err_msg = "data inconsistency"
        uid = obj.UID()
        brain = uuidToCatalogBrain(uid)
        catalog_data = self.get_catalog_indexdata(obj)

        # reference number obtained through the adapter is generated
        # dynamically, hence it should always be correct.
        # reference number in the catalog and in the metadata should match it.
        refnum = IReferenceNumber(obj).get_number()
        self.assertEqual(uid, refnum, brain.reference, err_msg)
        self.assertEqual(uid, refnum, catalog_data['reference'], err_msg)

        self.assertEqual(uid, brain.Description, obj.Description(), err_msg)

        self.assertEqual(uid, brain.getPath(), obj.absolute_url_path(), err_msg)
        self.assertEqual(uid, catalog_data['path'], obj.absolute_url_path(), err_msg)

        if not obj.portal_type == 'opengever.repository.repositoryfolder':
            return
        self.assertEqual(uid, brain.title_de, obj.get_prefixed_title_de(), err_msg)
        self.assertEqual(uid, brain.title_fr, obj.get_prefixed_title_fr(), err_msg)
        self.assertEqual(uid, catalog_data['sortable_title'], sortable_title(obj)(), err_msg)

    def assertEqual(self, uid, first, second, msg='not equal'):
        """Tests whether first and second are equal as determined by the '=='
        operator. If not, adds error to self.validation_errors, set
        self.validation_failed to true and log the error.
        """
        if not first == second:
            self.validation_errors[uid].append((first, second, msg))
            self.validation_failed = True
            logger.error(u"{}: {} ({}, {})".format(uid, msg, first, second))

    def get_catalog_indexdata(self, obj):
        """Return the catalog index data for an object as dict.
        """
        rid = self.catalog.getrid('/'.join(obj.getPhysicalPath()))
        return self.catalog.getIndexDataForRID(rid)


class FakeOptions(object):
    dry_run = False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', dest='site_root', default=None,
                        help='Absolute path to the Plone site')
    parser.add_argument('-m', dest='mapping', default=None,
                        help='Path to the mapping xlsx')
    parser.add_argument('-o', dest='output', default=None,
                        help='Path to the output xlsx')
    options = parser.parse_args(sys.argv[3:])
    app = setup_app()

    setup_plone(app, options)

    analyser = RepositoryExcelAnalyser(options.mapping, options.output)
    analyser.analyse()
    analyser.export_to_excel()


if __name__ == '__main__':
    main()
